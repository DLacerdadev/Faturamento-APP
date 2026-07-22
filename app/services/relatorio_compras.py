"""
Serviço de montagem do Relatório de Compras/Entregas por funcionário
(layout "zip").

VISÃO GERAL
-----------
A partir de um conjunto de PEDIDOS de compra selecionados na tela
(/relatorio-compras), este módulo produz uma planilha Excel onde CADA FUNCIONÁRIO
recebe uma ou mais linhas. As colunas são organizadas em 5 CATEGORIAS de itens:

    Exame | EPI | Uniforme | Equipamento | Treinamento

Cada categoria ocupa duas colunas na planilha: o NOME do item e a QUANTIDADE.

As fontes de dados são:
    - EpiPurchasePackage / EpiPurchaseItem  -> categorias 'epi' | 'uniforme' | 'equipamento'
      (a categoria vem do PACOTE; o item traz employee_numcad, employee_nome,
       descricao, quantidade)
    - MedicalExam                            -> categoria 'exame'
      (casa por matrícula/numcad; cada exame do catálogo com valor > 0 vira uma
       entrada com o nome amigável do exame)
    - TrainingRecord                         -> categoria 'treinamento'
      (employee_numcad, treinamento_nome, quantidade)

O "ZIP" (a parte mais delicada — ver montar_linhas_zip):
    Para cada funcionário, cada categoria tem uma LISTA de entradas
    (item + quantidade somada). Geramos N linhas por funcionário, onde
    N = maior número de entradas entre as 5 categorias. Na linha i colocamos
    a i-ésima entrada de cada categoria; se uma categoria tem menos que i
    entradas, as células dela ficam VAZIAS naquela linha. Matrícula, nome e
    centro de custo repetem em todas as linhas do funcionário.

REGRAS DE AGRUPAMENTO POR ITEM (dentro de cada categoria, por funcionário):
    - itens com nome DIFERENTE  -> entradas separadas
    - itens com nome IGUAL      -> UMA entrada, com a quantidade SOMADA
"""

from typing import Any, Dict, List, Optional, Tuple
from datetime import date, datetime
import logging

import pandas as pd

logger = logging.getLogger(__name__)

# Ordem/nome das colunas do Excel (contrato fixo pedido na tarefa).
COLUNAS_RELATORIO: List[str] = [
    "Centro de Custo", "Matrícula", "Funcionário",
    "Exame", "Exame Qtd",
    "EPI", "EPI Qtd",
    "Uniforme", "Uniforme Qtd",
    "Equipamento", "Equip Qtd",
    "Treinamento", "Treino Qtd",
]

# As 5 categorias, cada uma com (coluna do nome, coluna da quantidade).
# A ordem aqui define a ordem das entradas no "zip".
CATEGORIAS: List[Tuple[str, str, str]] = [
    ("exame", "Exame", "Exame Qtd"),
    ("epi", "EPI", "EPI Qtd"),
    ("uniforme", "Uniforme", "Uniforme Qtd"),
    ("equipamento", "Equipamento", "Equip Qtd"),
    ("treinamento", "Treinamento", "Treino Qtd"),
]

# Mapa: campo (coluna) do MedicalExam -> rótulo amigável do exame. Cada campo
# com valor > 0 no registro vira uma entrada de exame para o funcionário.
EXAM_FIELD_LABELS: List[Tuple[str, str]] = [
    ("clinic", "Clínico"),
    ("audio", "Audiometria"),
    ("acuid", "Acuidade Visual"),
    ("hemo", "Hemograma"),
    ("lipidograma", "Lipidograma"),
    ("rx_coluna", "RX Coluna"),
    ("met_e_cet", "Metahemo/Carboxi"),
    ("acet_u", "Acetona Urina"),
    ("hg", "Hemoglobina"),
    ("retic", "Reticulócitos"),
    ("ac_trans", "Ácido Trans-mucônico"),
    ("eeg", "EEG"),
    ("ecg", "ECG"),
    ("etanol", "Etanol"),
    ("glice", "Glicemia"),
    ("gama_gt", "Gama GT"),
    ("tgp", "TGP"),
    ("rx_torax", "RX Tórax"),
    ("espiro", "Espirometria"),
    ("rx_lomb", "RX Lombar"),
    ("aval_psicossocial", "Aval. Psicossocial"),
]


# ----------------------------------------------------------------------------
# Coleta das entradas por funcionário
# ----------------------------------------------------------------------------

def _emp_key(numcad: Optional[int], nome: Optional[str]) -> Tuple[str, str]:
    """Chave estável do funcionário. Prioriza a matrícula (numcad); quando
    ausente, cai para o nome normalizado. Isso mantém junto tudo do mesmo
    funcionário mesmo que uma das fontes não tenha a matrícula."""
    if numcad is not None:
        return ("n", str(numcad))
    return ("s", (nome or "").strip().upper())


class _EmployeeBucket:
    """Acumula, para UM funcionário, as entradas de cada categoria.

    entradas[categoria] = { nome_item: quantidade_somada }
    Usa dict por nome para agrupar itens iguais somando a quantidade.
    A ordem de inserção do dict (Python 3.7+) preserva a ordem em que os
    itens apareceram — usada como ordem das entradas no zip.
    """

    def __init__(self, numcad: Optional[int], nome: str, codccu: str, nome_ccu: str):
        self.numcad = numcad
        self.nome = nome
        self.codccu = codccu
        self.nome_ccu = nome_ccu
        self.entradas: Dict[str, Dict[str, float]] = {
            cat: {} for cat, _, _ in CATEGORIAS
        }

    def preencher_identidade(self, numcad, nome, codccu, nome_ccu):
        """Completa campos de identidade quando alguma fonte tiver dados que
        as anteriores não tinham (ex.: numcad ou centro de custo)."""
        if self.numcad is None and numcad is not None:
            self.numcad = numcad
        if not self.nome and nome:
            self.nome = nome
        if not self.codccu and codccu:
            self.codccu = codccu
        if not self.nome_ccu and nome_ccu:
            self.nome_ccu = nome_ccu

    def add(self, categoria: str, nome_item: str, quantidade: float):
        nome_item = (nome_item or "").strip()
        if not nome_item:
            return
        try:
            q = float(quantidade or 0)
        except (TypeError, ValueError):
            q = 0.0
        bucket = self.entradas[categoria]
        bucket[nome_item] = bucket.get(nome_item, 0.0) + q


def _fmt_qtd(q: float):
    """Quantidade sem casas decimais quando inteira (2.0 -> 2)."""
    if q is None:
        return ""
    if float(q).is_integer():
        return int(q)
    return round(float(q), 2)


def coletar_por_funcionario(
    itens_compra: List[Dict[str, Any]],
    exames: List[Dict[str, Any]],
    treinamentos: List[Dict[str, Any]],
) -> Dict[Tuple[str, str], _EmployeeBucket]:
    """Junta as três fontes num dicionário {chave_func: _EmployeeBucket}.

    Espera dicts já "achatados" (montados pelo router a partir dos models),
    para manter este módulo testável sem sessão de banco:

    itens_compra: cada dict tem
        {categoria('epi'|'uniforme'|'equipamento'), numcad, nome,
         codccu, nome_ccu, descricao, quantidade}
    exames: cada dict tem
        {numcad, nome, codccu, nome_ccu, exames:[{nome, quantidade}, ...]}
    treinamentos: cada dict tem
        {numcad, nome, codccu, nome_ccu, treinamento_nome, quantidade}
    """
    buckets: Dict[Tuple[str, str], _EmployeeBucket] = {}

    def get_bucket(numcad, nome, codccu, nome_ccu) -> _EmployeeBucket:
        key = _emp_key(numcad, nome)
        b = buckets.get(key)
        if b is None:
            b = _EmployeeBucket(numcad, (nome or "").strip(), codccu or "", nome_ccu or "")
            buckets[key] = b
        else:
            b.preencher_identidade(numcad, nome, codccu, nome_ccu)
        return b

    # Compras (epi/uniforme/equipamento): categoria vem do pacote.
    for it in itens_compra:
        categoria = it.get("categoria")
        if categoria not in ("epi", "uniforme", "equipamento"):
            continue
        b = get_bucket(it.get("numcad"), it.get("nome"), it.get("codccu"), it.get("nome_ccu"))
        b.add(categoria, it.get("descricao"), it.get("quantidade"))

    # Exames: cada exame do catálogo (valor > 0) já veio como entrada nome+qtd.
    for ex in exames:
        b = get_bucket(ex.get("numcad"), ex.get("nome"), ex.get("codccu"), ex.get("nome_ccu"))
        for e in ex.get("exames", []):
            b.add("exame", e.get("nome"), e.get("quantidade", 1))

    # Treinamentos.
    for tr in treinamentos:
        b = get_bucket(tr.get("numcad"), tr.get("nome"), tr.get("codccu"), tr.get("nome_ccu"))
        b.add("treinamento", tr.get("treinamento_nome"), tr.get("quantidade", 1))

    return buckets


# ----------------------------------------------------------------------------
# ZIP: transforma cada funcionário em N linhas
# ----------------------------------------------------------------------------

def montar_linhas_zip(buckets: Dict[Tuple[str, str], _EmployeeBucket]) -> List[Dict[str, Any]]:
    """Aplica o "zip" e devolve a lista de linhas (dicts prontos p/ o Excel).

    Para cada funcionário:
      1. Converte o dict {nome_item: qtd} de cada categoria numa LISTA ordenada
         de tuplas (nome_item, qtd) — a ordem é a de inserção (ordem em que os
         itens apareceram nas fontes).
      2. N = max(len(lista) das 5 categorias). Se N == 0 (funcionário sem nada),
         gera 1 linha só com a identidade (não deveria ocorrer, mas é defensivo).
      3. Para i em 0..N-1: monta uma linha com a identidade repetida e, para
         cada categoria, coloca a i-ésima (nome, qtd) se existir; senão célula
         vazia.

    Retorna as linhas ordenadas por centro de custo, depois nome do funcionário,
    para uma leitura estável da planilha.
    """
    linhas: List[Dict[str, Any]] = []

    # Ordena funcionários por CC e nome para a saída ficar estável/legível.
    ordenados = sorted(
        buckets.values(),
        key=lambda b: ((b.codccu or ""), (b.nome or "").upper(), b.numcad or 0),
    )

    for b in ordenados:
        # (1) listas ordenadas de entradas por categoria
        listas: Dict[str, List[Tuple[str, float]]] = {}
        for cat, _, _ in CATEGORIAS:
            listas[cat] = [(nome, qtd) for nome, qtd in b.entradas[cat].items()]

        # (2) N = maior número de entradas entre as categorias
        n = max((len(listas[cat]) for cat, _, _ in CATEGORIAS), default=0)
        cc_label = b.codccu or ""
        if b.nome_ccu:
            cc_label = f"{b.codccu} — {b.nome_ccu}" if b.codccu else b.nome_ccu

        if n == 0:
            # Funcionário sem nenhuma entrada: 1 linha só com identidade.
            linhas.append(_linha_base(cc_label, b))
            continue

        # (3) zip linha a linha
        for i in range(n):
            linha = _linha_base(cc_label, b)
            for cat, col_nome, col_qtd in CATEGORIAS:
                lst = listas[cat]
                if i < len(lst):
                    nome_item, qtd = lst[i]
                    linha[col_nome] = nome_item
                    linha[col_qtd] = _fmt_qtd(qtd)
                # else: deixa vazio (já é "" pelo _linha_base)
            linhas.append(linha)

    return linhas


def _linha_base(cc_label: str, b: _EmployeeBucket) -> Dict[str, Any]:
    """Linha com identidade preenchida e todas as colunas de item vazias."""
    linha = {col: "" for col in COLUNAS_RELATORIO}
    linha["Centro de Custo"] = cc_label
    linha["Matrícula"] = b.numcad if b.numcad is not None else ""
    linha["Funcionário"] = b.nome or ""
    return linha


# ----------------------------------------------------------------------------
# Excel
# ----------------------------------------------------------------------------

def montar_relatorio(
    itens_compra: List[Dict[str, Any]],
    exames: List[Dict[str, Any]],
    treinamentos: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    """Pipeline completo (coleta -> zip). Retorna as linhas do relatório."""
    buckets = coletar_por_funcionario(itens_compra, exames, treinamentos)
    return montar_linhas_zip(buckets)


def relatorio_to_excel_bytes(linhas: List[Dict[str, Any]]) -> bytes:
    """Gera os bytes do .xlsx a partir das linhas já montadas.

    Segue o padrão do projeto (pandas + openpyxl, autoajuste de largura). As
    colunas seguem exatamente COLUNAS_RELATORIO, na ordem contratada.
    """
    from io import BytesIO
    from openpyxl.styles import PatternFill, Font, Alignment

    if not linhas:
        df = pd.DataFrame(columns=COLUNAS_RELATORIO)
    else:
        df = pd.DataFrame(linhas)
        # Garante todas as colunas na ordem correta (defensivo).
        for c in COLUNAS_RELATORIO:
            if c not in df.columns:
                df[c] = ""
        df = df[COLUNAS_RELATORIO]

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        sheet_name = "Compras por Funcionário"[:31]
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]

        # Cabeçalho no tema institucional (navy + âmbar), igual à Folha Senior.
        header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
        header_font = Font(bold=True, color="D4A84B")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for idx, col in enumerate(df.columns):
            if idx < 26:
                col_letter = chr(65 + idx)
            else:
                col_letter = chr(64 + idx // 26) + chr(65 + idx % 26)
            max_length = max(
                int(df[col].fillna("").astype(str).apply(len).max()) if len(df) > 0 else 0,
                len(str(col)),
            ) + 2
            ws.column_dimensions[col_letter].width = min(max_length, 40)

    output.seek(0)
    return output.getvalue()


def gerar_nome_arquivo(data_ini: Optional[str], data_fim: Optional[str], codccu: Optional[str]) -> str:
    """Nome padronizado do arquivo do relatório de compras."""
    def fmt(d):
        if not d:
            return ""
        return str(d).replace("-", "")[:8]
    ini = fmt(data_ini)
    fim = fmt(data_fim)
    ccu = (codccu or "todos").replace(" ", "_").replace("/", "_")
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    periodo = f"{ini}-{fim}" if (ini or fim) else "periodo"
    return f"relatorio_compras_{ccu}_{periodo}_{ts}.xlsx"
