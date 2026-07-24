"""Serviço de conciliação contábil (Etapa 3 do Plano de Execução).

Monta a "ponte" entre dois recortes de valores de uma competência:

    competência inteira  =  recorte mensal  +  fora do recorte  +  (não classificados)

- competência inteira: soma de TODOS os códigos de cálculo (CODCAL) da folha.
- recorte mensal: soma dos CODCAL classificados como 'entra no recorte mensal'
  (o que a contabilidade confere no relatório mensal da Senior).
- fora do recorte: soma dos CODCAL classificados como 'fora'.
- não classificados: CODCAL presentes na folha sem classificação cadastrada —
  enquanto existirem, a conciliação fica 'incompleta' (nunca fecha no silêncio).

`montar_conciliacao` é uma função PURA sobre a lista já buscada do WS
(fetch_payroll) — não faz I/O. Isso a torna o alvo natural dos testes da
Etapa 4 (spec 005).

Nenhum dado individual de funcionário atravessa para o resultado: a agregação
é sempre por CODCAL -> evento (valor total + quantidade de lançamentos).
"""
from typing import Any, Dict, List, Optional

# Tolerância de arredondamento para considerar a ponte "fechada" (centavos).
_EPS = 0.005


def montar_conciliacao(
    payroll_rows: List[Dict[str, Any]],
    classificacoes: Dict[int, Dict[str, Any]],
) -> Dict[str, Any]:
    """Agrega as linhas da folha por CODCAL -> evento e monta a conciliação.

    Args:
        payroll_rows: lista plana de fetch_payroll (cada item = uma linha de
            evento de um funcionário, com 'codcal', 'codigo_evento',
            'descricao_evento', 'valor_evento').
        classificacoes: mapa codcal(int) -> {'recorte_mensal': bool,
            'descricao': str|None, 'origem': str}.

    Returns:
        dict com 'totais', 'codcals' (decomposição), 'nao_classificados' e
        'status' — ver specs/004-relatorio-conciliacao/data-model.md.
    """
    # Agregação por codcal separando PROVENTOS e DESCONTOS pelo tipo do evento
    # (tipo_evento: 1,2 = provento/vantagem; 3 = desconto; demais = bases de FGTS/
    # INSS etc., que NÃO entram no resumo da folha). Assim os totais batem com o
    # "Resumo dos Valores Totais" da folha (Proventos / Descontos / Líquido), em
    # vez de somar tudo junto (que não corresponde a nenhuma linha da folha).
    por_codcal: Dict[int, Dict[str, Any]] = {}

    for row in payroll_rows:
        codcal = _safe_int(row.get("codcal"))
        if codcal is None:
            codcal = 0  # linhas sem codcal caem num balde "0" (aparecerá como não classificado)
        valor = abs(_safe_float(row.get("valor_evento")))
        tipe = _safe_int(row.get("tipo_evento")) or 0
        cod_ev = row.get("codigo_evento")
        cod_ev_key = str(cod_ev) if cod_ev is not None else "?"

        bucket = por_codcal.setdefault(codcal, {"proventos": 0.0, "descontos": 0.0, "eventos": {}})
        ev = bucket["eventos"].setdefault(
            cod_ev_key,
            {"codigo": cod_ev_key, "descricao": row.get("descricao_evento"),
             "proventos": 0.0, "descontos": 0.0, "lancamentos": 0},
        )
        if tipe == 3:
            bucket["descontos"] += valor
            ev["descontos"] += valor
        elif tipe in (1, 2):
            bucket["proventos"] += valor
            ev["proventos"] += valor
        # demais tipos (bases): não somam em proventos/descontos
        ev["lancamentos"] += 1
        if not ev["descricao"] and row.get("descricao_evento"):
            ev["descricao"] = row.get("descricao_evento")

    codcals_out: List[Dict[str, Any]] = []
    nao_classificados: List[int] = []
    prov_inteira = desc_inteira = 0.0
    prov_mensal = prov_fora = 0.0

    for codcal in sorted(por_codcal.keys()):
        bucket = por_codcal[codcal]
        proventos = round(bucket["proventos"], 2)
        descontos = round(bucket["descontos"], 2)
        liquido = round(proventos - descontos, 2)
        prov_inteira += proventos
        desc_inteira += descontos

        cls = classificacoes.get(codcal)
        if cls is None:
            classificacao = "nao_classificado"
            descricao = None
            origem = None
            nao_classificados.append(codcal)
        elif cls.get("recorte_mensal"):
            classificacao = "mensal"
            descricao = cls.get("descricao")
            origem = cls.get("origem") or "manual"
            prov_mensal += proventos
        else:
            classificacao = "fora"
            descricao = cls.get("descricao")
            origem = cls.get("origem") or "manual"
            prov_fora += proventos

        eventos = sorted(
            (
                {
                    "codigo": e["codigo"],
                    "descricao": e["descricao"],
                    "proventos": round(e["proventos"], 2),
                    "descontos": round(e["descontos"], 2),
                    "valor_total": round(e["proventos"] - e["descontos"], 2),
                    "lancamentos": e["lancamentos"],
                }
                for e in bucket["eventos"].values()
            ),
            key=lambda e: e["codigo"],
        )

        codcals_out.append({
            "codcal": codcal,
            "descricao": descricao,
            "classificacao": classificacao,
            "origem": origem,
            "proventos": proventos,
            "descontos": descontos,
            "liquido": liquido,
            "valor_total": proventos,   # a PONTE do faturamento é sobre proventos
            "eventos": eventos,
        })

    prov_inteira = round(prov_inteira, 2)
    desc_inteira = round(desc_inteira, 2)
    prov_mensal = round(prov_mensal, 2)
    prov_fora = round(prov_fora, 2)
    residuo = round(prov_inteira - prov_mensal - prov_fora, 2)

    if nao_classificados:
        status = "incompleta"
    elif abs(residuo) > _EPS:
        status = "com_residuo"
    else:
        status = "fechada"

    return {
        "status": status,
        "totais": {
            # Ponte do faturamento (base = proventos)
            "competencia_inteira": prov_inteira,
            "recorte_mensal": prov_mensal,
            "fora_recorte": prov_fora,
            "residuo": residuo,
            # Conferência com o resumo da folha (Proventos / Descontos / Líquido)
            "proventos": prov_inteira,
            "descontos": desc_inteira,
            "liquido": round(prov_inteira - desc_inteira, 2),
        },
        "codcals": codcals_out,
        "nao_classificados": nao_classificados,
    }


def sugerir_classificacao(eventos: List[Dict[str, Any]]) -> Optional[bool]:
    """Heurística leve de sugestão (NÃO grava): a partir dos eventos agregados
    de um codcal não classificado, sugere True (provável recorte mensal) quando
    aparecem eventos típicos de folha mensal. Retorna None quando não há palpite.

    O gestor sempre confirma antes de gravar (SC-3) — isto só pré-preenche a UI.
    """
    marcadores_mensal = ("SALARIO", "SALÁRIO", "HORAS", "ADICIONAL", "DSR", "REPOUSO")
    for ev in eventos:
        desc = (ev.get("descricao") or "").upper()
        if any(m in desc for m in marcadores_mensal):
            return True
    return None


def conciliacao_para_xlsx(resultado: Dict[str, Any]) -> bytes:
    """Converte o resultado da conciliação em uma planilha .xlsx (bytes).

    3 abas: Resumo, Decomposição (por codcal), Eventos (codcal x evento).
    Sem nenhum dado individual de funcionário.
    """
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    bold = Font(bold=True)

    # --- Aba Resumo ---
    ws = wb.active
    ws.title = "Resumo"
    totais = resultado.get("totais", {})
    linhas_resumo = [
        ("Competência", resultado.get("periodo")),
        ("Centro de custo", resultado.get("codccu") or "Todos"),
        ("Nome do CCU", resultado.get("nomccu") or ""),
        ("Gerado em", resultado.get("gerado_em")),
        ("Gerado por", resultado.get("gerado_por")),
        ("Status", resultado.get("status")),
        ("", ""),
        ("PONTE DO FATURAMENTO (base = proventos)", ""),
        ("Competência inteira", totais.get("competencia_inteira")),
        ("Recorte mensal", totais.get("recorte_mensal")),
        ("Fora do recorte", totais.get("fora_recorte")),
        ("Resíduo", totais.get("residuo")),
        ("", ""),
        ("CONFERÊNCIA COM O RESUMO DA FOLHA", ""),
        ("Proventos", totais.get("proventos")),
        ("Descontos", totais.get("descontos")),
        ("Líquido", totais.get("liquido")),
    ]
    for i, (rot, val) in enumerate(linhas_resumo, start=1):
        ws.cell(row=i, column=1, value=rot).font = bold
        ws.cell(row=i, column=2, value=val)
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 40

    # --- Aba Decomposição ---
    wd = wb.create_sheet("Decomposição")
    headers = ["CODCAL", "Descrição", "Classificação", "Origem", "Proventos", "Descontos", "Líquido"]
    for c, h in enumerate(headers, start=1):
        wd.cell(row=1, column=c, value=h).font = bold
    for r, cc in enumerate(resultado.get("codcals", []), start=2):
        wd.cell(row=r, column=1, value=cc.get("codcal"))
        wd.cell(row=r, column=2, value=cc.get("descricao"))
        wd.cell(row=r, column=3, value=cc.get("classificacao"))
        wd.cell(row=r, column=4, value=cc.get("origem"))
        wd.cell(row=r, column=5, value=cc.get("proventos"))
        wd.cell(row=r, column=6, value=cc.get("descontos"))
        wd.cell(row=r, column=7, value=cc.get("liquido"))
    for col, w in zip("ABCDEFG", (10, 32, 16, 12, 15, 15, 15)):
        wd.column_dimensions[col].width = w

    # --- Aba Eventos ---
    we = wb.create_sheet("Eventos")
    headers = ["CODCAL", "Evento", "Descrição", "Proventos", "Descontos", "Lançamentos"]
    for c, h in enumerate(headers, start=1):
        we.cell(row=1, column=c, value=h).font = bold
    r = 2
    for cc in resultado.get("codcals", []):
        for ev in cc.get("eventos", []):
            we.cell(row=r, column=1, value=cc.get("codcal"))
            we.cell(row=r, column=2, value=ev.get("codigo"))
            we.cell(row=r, column=3, value=ev.get("descricao"))
            we.cell(row=r, column=4, value=ev.get("proventos"))
            we.cell(row=r, column=5, value=ev.get("descontos"))
            we.cell(row=r, column=6, value=ev.get("lancamentos"))
            r += 1
    for col, w in zip("ABCDEF", (10, 12, 34, 15, 15, 14)):
        we.column_dimensions[col].width = w

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _safe_int(v) -> Optional[int]:
    try:
        if v is None or v == "":
            return None
        return int(v)
    except (ValueError, TypeError):
        return None


def _safe_float(v) -> float:
    try:
        if v is None or v == "":
            return 0.0
        return float(v)
    except (ValueError, TypeError):
        return 0.0
