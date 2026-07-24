"""
Serviço de exportação para Excel.
Converte dados de faturamento em arquivos Excel para download.
"""
from typing import List, Dict, Any, Optional, Union


def _safe_salario_cell(val: Any) -> float:
    if val is None:
        return 0.0
    try:
        return float(val)
    except (TypeError, ValueError):
        return 0.0
from io import BytesIO
import pandas as pd
import logging
import re
from datetime import datetime
from openpyxl.styles import PatternFill, Font
from app.services.billing_processor import calcular_totais_remuneracao

logger = logging.getLogger(__name__)


def invoice_to_excel_bytes(
    invoice_rows: List[Dict[str, Any]], 
    filename_hint: str = "fatura"
) -> bytes:
    """
    Converte lista de dicts de fatura em bytes de arquivo Excel.
    
    Args:
        invoice_rows: Lista de dicts com dados da fatura
        filename_hint: Nome sugerido para o arquivo (não usado internamente)
        
    Returns:
        Bytes do arquivo Excel pronto para download
    """
    if not invoice_rows:
        df = pd.DataFrame({"Mensagem": ["Nenhum dado encontrado para os filtros selecionados"]})
    else:
        df = pd.DataFrame(invoice_rows)
    
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine="openpyxl") as writer:  # type: ignore
        df.to_excel(writer, index=False, sheet_name="Faturamento")
        
        worksheet = writer.sheets["Faturamento"]
        for idx, col in enumerate(df.columns):
            if idx < 26:
                max_length = max(
                    int(df[col].fillna('').astype(str).apply(len).max()) if len(df) > 0 else 0,
                    len(str(col))
                ) + 2
                worksheet.column_dimensions[chr(65 + idx)].width = min(max_length, 50)
    
    output.seek(0)
    return output.getvalue()


def invoice_to_excel_multi_sheet(
    data_sheets: Dict[str, List[Dict[str, Any]]],
    filename_hint: str = "fatura"
) -> bytes:
    """
    Cria Excel com múltiplas abas.
    
    Args:
        data_sheets: Dict onde chave é nome da aba e valor é lista de dados
        filename_hint: Nome sugerido para o arquivo
        
    Returns:
        Bytes do arquivo Excel
    """
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine="openpyxl") as writer:  # type: ignore
        for sheet_name, rows in data_sheets.items():
            if not rows:
                df = pd.DataFrame({"Mensagem": ["Nenhum dado"]})
            else:
                df = pd.DataFrame(rows)
            
            safe_name = sheet_name[:31]
            df.to_excel(writer, index=False, sheet_name=safe_name)
            
            worksheet = writer.sheets[safe_name]
            for idx, col in enumerate(df.columns):
                if idx < 26:
                    max_length = max(
                        int(df[col].fillna('').astype(str).apply(len).max()) if len(df) > 0 else 0,
                        len(str(col))
                    ) + 2
                    worksheet.column_dimensions[chr(65 + idx)].width = min(max_length, 50)
    
    output.seek(0)
    return output.getvalue()


def generate_invoice_filename(
    periodo: str, 
    codccu: Optional[str] = None, 
    prefix: str = "fatura_telos"
) -> str:
    """
    Gera nome de arquivo padronizado para fatura.
    
    Args:
        periodo: Período no formato YYYY-MM-DD
        codccu: Código do centro de custo (opcional)
        prefix: Prefixo do nome do arquivo
        
    Returns:
        Nome do arquivo formatado
    """
    periodo_fmt = periodo.replace("-", "")[:6]
    ccu_part = codccu.replace(" ", "_").replace("/", "_") if codccu else "todos"
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    return f"{prefix}_{periodo_fmt}_{ccu_part}_{timestamp}.xlsx"


def payroll_to_excel_bytes(
    grouped_employees: List[Dict[str, Any]],
    periodo: str,
    codccu: str
) -> bytes:
    """
    Converte dados de folha de pagamento agrupados por funcionário em Excel.
    Gera uma planilha com uma linha por funcionário e eventos como colunas.
    Cada evento gera duas colunas: (ref) e (valor).
    
    Args:
        grouped_employees: Lista de funcionários com seus eventos
        periodo: Período de referência
        codccu: Centro de custo
        
    Returns:
        Bytes do arquivo Excel
    """
    all_event_types: Dict[str, str] = {}
    for emp in grouped_employees:
        for ev in emp.get("eventos", []):
            cod = ev.get("codigo_evento", "")
            desc = ev.get("descricao_evento", "")
            if cod and cod not in all_event_types:
                all_event_types[cod] = desc.strip() if desc else f"Evento {cod}"
    
    sorted_events = sorted(all_event_types.items(), key=lambda x: x[0])
    
    def format_date(date_val):
        """Formata data para dd/MM/yyyy"""
        if not date_val:
            return ""
        if isinstance(date_val, str):
            for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%Y-%m-%dT%H:%M:%S"]:
                try:
                    dt = datetime.strptime(date_val[:10], fmt)
                    return dt.strftime("%d/%m/%Y")
                except (ValueError, TypeError):
                    continue
            return date_val
        if hasattr(date_val, 'strftime'):
            return date_val.strftime("%d/%m/%Y")
        return str(date_val)
    
    rows = []
    for emp in grouped_employees:
        row: Dict[str, Any] = {
            "Matricula": emp.get("matricula", ""),
            "Nome": emp.get("nome_funcionario", ""),
            "Cargo": emp.get("cargo", ""),
            "Salario Base": _safe_salario_cell(emp.get("salario")),
            "Admissao": format_date(emp.get("data_admissao", "")),
            "Demissao": format_date(emp.get("data_afastamento", "")),
        }
        
        event_map: Dict[str, Dict[str, float]] = {}
        for ev in emp.get("eventos", []):
            cod = ev.get("codigo_evento", "")
            if cod:
                event_map[cod] = {
                    "ref": ev.get("referencia_evento", 0) or 0,
                    "valor": ev.get("valor_evento", 0) or 0
                }
        
        for cod, desc in sorted_events:
            ref_col = f"{desc} (ref)"
            val_col = f"{desc} (valor)"
            if cod in event_map:
                row[ref_col] = event_map[cod]["ref"]
                row[val_col] = event_map[cod]["valor"]
            else:
                row[ref_col] = 0
                row[val_col] = 0
        
        rows.append(row)
    
    if not rows:
        df = pd.DataFrame({"Mensagem": ["Nenhum dado encontrado"]})
    else:
        df = pd.DataFrame(rows)
    
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        sheet_name = f"Folha {codccu}"[:31]
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        
        worksheet = writer.sheets[sheet_name]
        for idx, col in enumerate(df.columns):
            if idx < 26:
                col_letter = chr(65 + idx)
            elif idx < 702:
                col_letter = chr(64 + idx // 26) + chr(65 + idx % 26)
            else:
                col_letter = "A"
            max_length = max(
                int(df[col].fillna('').astype(str).apply(len).max()) if len(df) > 0 else 0,
                len(str(col))
            ) + 2
            worksheet.column_dimensions[col_letter].width = min(max_length, 30)
    
    output.seek(0)
    return output.getvalue()


def generate_payroll_filename(periodo: str, codccu: str) -> str:
    """
    Gera nome de arquivo para folha de pagamento.
    """
    periodo_fmt = periodo.replace("-", "")[:6]
    ccu_safe = codccu.replace(" ", "_").replace("/", "_")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"folha_telos_{periodo_fmt}_{ccu_safe}_{timestamp}.xlsx"


FEMSA_COLUMNS = [
    "Empresa", "Mês Referência", "Nº Posicão", "CNPJ (Inserir CNPJ agência)", "CNPJ FEMSA",
    "Nome", "CPF", "Função", "Unidade - ", "Centro de Custo - Femsa", "Cargo - Femsa",
    "Motivo - Femsa", "Salário", "Dt Admissão", "Dt Demissão", "Término Ctr.",
    "Período afastamento", "SALARIO DIA (Qtde)", "SALARIO DIA (Valor)", "HORAS EXTRAS (Qtde)",
    "HORAS EXTRAS (Valor)", "ADICIONAL NOTURNO (Qtde)", "ADICIONAL NOTURNO (Valor)",
    "ADIC. PERICULOSIDADE", "REMUNERACAO VARIAVEL MENSAL", "SALÁRIO BRUTO ", "SALÁRIO LÍQUIDO",
    "Refeitório - SIM/NÃO", "Período Benefício", "PAGTO. VALE-TRANSPORTE (Qtde)",
    "PAGTO. VALE-TRANSPORTE (Valor)", "PAGTO. VALE REFEICAO (Qtde)", "PAGTO. VALE REFEICAO (Valor)",
    "AJUDA CUSTO COMBUSTÍVEL / KM", "PREMIO/BONUS", "VALE TRANSPORTE NAO UTILIZADO",
    "REEMB. VALE REFEICAO INDEVIDO/DEVOLVIDO", "REEMB. DESPESAS KM/ESTAC/PEDAGIO",
    "REEMB. DESC.DE FALTAS/ATRASOS/D.S.R. (HS) (Valor)", "LICENCA PATERNIDADE (Valor)",
    "ATESTADO MEDICO DIA (Valor)", "AUXILIO DOENÇA (Valor)",
    "D.S.R. INTEGRACAO S/ ADICIONAL NOTURNO", "D.S.R. INTEGRACAO S/ HORA EXTRA",
    "D.S.R. INTEGRACAO S/ VARIAVEL/PREMIO", "ADICIONAL NOTURNO  - MES ANTERIOR (Valor)",
    "VARIAVEL MÊS ANTERIOR", "HORA EXTRA - MES ANTERIOR (valor)", "INTERJORNADA MES ANTERIOR (Valor)",
    "ADIANTAMENTO QUINZENAL", "DIFERENCA DE SALARIO", "DESCONTO DE ADIANTAMENTO",
    "DESCONTO DE QUEBRA DE CAIXA", "FALTAS E ATRASOS/DSR - MES ANT. (Valor)",
    "REEMB. D.S.R. S/FALTAS (DIA) (Valor)", "SALDO NEGATIVO", "DESC. SALDO NEGATIVO",
    "SALDO SALARIO DIA RESCISAO (Valor)", "Pensão Judicial",
    "Total Remuneração", "Encargos Sociais",
    "(FAT) EXAMES MEDICOS",
    "EXAMES MEDICOS COMPLEMENTARES", "SEGURO DE VIDA", "TAXA EXAMES MEDICOS(Valor)",
    "TAXA EXAMES MEDICOS(%)", "TAXA EXAMES MEDICOS COMPLEMENTARES (Valor)",
    "TAXA EXAMES MEDICOS COMPLEMENTARES (%)", "ENCARGOS DE FOLHA", "TAXA FATURAMENTO (VALOR)",
    "TAXA FATURAMENTO (%)", "TAXA CONTRATO (VALOR)", "TAXA CONTRATO (%)", "TRIBUTOS (VALOR)",
    "TRIBUTOS (%)", "ENCARGOS (VALOR)", "ENCARGOS (%)", "Sub-Total", "Total Geral"
]

# Colunas NOVAS (Frente 3 preenche os valores por funcionário). Só aparecem na
# planilha quando o modelo de faturamento as inclui na sua lista de colunas.
# Entram no Sub-Total (custo) via calcular_faturamento quando existirem no df.
EXTRA_COST_COLUMNS = [
    "UNIFORMES (Valor)",
    "EPIS (Valor)",
    "EQUIPAMENTOS (Valor)",
    "TREINAMENTOS (Valor)",
]

# Conjunto de TODAS as colunas conhecidas (base para montar/editar modelos =
# GERAL). FEMSA idêntico ao de hoje; as 4 colunas de itens entram logo APÓS
# "SEGURO DE VIDA" (não no fim da planilha).
_POS_APOS_SEGURO = FEMSA_COLUMNS.index("SEGURO DE VIDA") + 1
GERAL_COLUMNS = (
    FEMSA_COLUMNS[:_POS_APOS_SEGURO]
    + EXTRA_COST_COLUMNS
    + FEMSA_COLUMNS[_POS_APOS_SEGURO:]
)

EVENT_TO_FEMSA_MAPPING = {
    7: (None, "AUXILIO DOENÇA (Valor)"),
    13: (None, "LICENCA PATERNIDADE (Valor)"),
    29: (None, "INTERJORNADA MES ANTERIOR (Valor)"),
    106: (None, "AUXILIO DOENÇA (Valor)"),
    200: ("SALARIO DIA (Qtde)", "SALARIO DIA (Valor)"),
    202: (None, "REEMB. DESC.DE FALTAS/ATRASOS/D.S.R. (HS) (Valor)"),
    206: (None, "AUXILIO DOENÇA (Valor)"),
    212: (None, "LICENCA PATERNIDADE (Valor)"),
    213: (None, "ATESTADO MEDICO DIA (Valor)"),
    257: ("HORAS EXTRAS (Qtde)", "HORAS EXTRAS (Valor)"),
    265: (None, "D.S.R. INTEGRACAO S/ HORA EXTRA"),
    652: None,
    656: (None, "ADICIONAL NOTURNO  - MES ANTERIOR (Valor)"),
    850: None,
    851: None,
    1550: (None, "SALDO SALARIO DIA RESCISAO (Valor)"),
    1701: (None, "SALDO NEGATIVO"),
    1950: (None, "ADICIONAL NOTURNO (Valor)"),
    1957: (None, "DESCONTO DE QUEBRA DE CAIXA"),
    1961: None,
    1962: (None, "DIFERENCA DE SALARIO"),
    1975: None,
    2000: None,
    2004: None,
    2250: (None, "Pensão Judicial"),
    2251: (None, "Pensão Judicial"),
    2252: (None, "Pensão Judicial"),
    2255: (None, "Pensão Judicial"),
    2463: None,
    2464: None,
    2469: None,
    2470: None,
    2500: None,
    3029: ("ADICIONAL NOTURNO (Qtde)", "ADICIONAL NOTURNO (Valor)"),
    3030: (None, "D.S.R. INTEGRACAO S/ ADICIONAL NOTURNO"),
    3031: ("PAGTO. VALE REFEICAO (Qtde)", "PAGTO. VALE REFEICAO (Valor)"),
    3035: (None, "FALTAS E ATRASOS/DSR - MES ANT. (Valor)"),
    3081: (None, "REEMB. DESC.DE FALTAS/ATRASOS/D.S.R. (HS) (Valor)"),
    3084: (None, "REEMB. D.S.R. S/FALTAS (DIA) (Valor)"),
    3127: None,
    3130: None,
    3158: (None, "VALE TRANSPORTE NAO UTILIZADO"),
    3161: (None, "DIFERENCA DE SALARIO"),
    3199: (None, "Pensão Judicial"),
    3227: None,
    3231: (None, "DESC. SALDO NEGATIVO"),
    3233: (None, "ADICIONAL NOTURNO (Valor)"),
    5504: (None, "DESCONTO DE QUEBRA DE CAIXA"),
    3650: ("ADICIONAL NOTURNO (Qtde)", "ADICIONAL NOTURNO (Valor)"),
    5606: (None, "REMUNERACAO VARIAVEL MENSAL"),
    5996: (None, "REMUNERACAO VARIAVEL MENSAL"),
    3000: (None, "HORA EXTRA - MES ANTERIOR (valor)"),
    3003: (None, "HORA EXTRA - MES ANTERIOR (valor)"),
    3019: (None, "HORA EXTRA - MES ANTERIOR (valor)"),
    3073: (None, "HORA EXTRA - MES ANTERIOR (valor)"),
    3119: (None, "HORA EXTRA - MES ANTERIOR (valor)"),
    3148: (None, "HORA EXTRA - MES ANTERIOR (valor)"),
    3174: (None, "HORA EXTRA - MES ANTERIOR (valor)"),
    3186: (None, "HORA EXTRA - MES ANTERIOR (valor)"),
    3216: (None, "HORA EXTRA - MES ANTERIOR (valor)"),
    3246: (None, "HORA EXTRA - MES ANTERIOR (valor)"),
    3250: (None, "HORA EXTRA - MES ANTERIOR (valor)"),
    3615: (None, "HORA EXTRA - MES ANTERIOR (valor)"),
    4258: (None, "HORA EXTRA - MES ANTERIOR (valor)"),
    # Códigos só usados em DEV (_fetch_payroll_local) para alinhar PayrollItemType ao grid FEMSA
    9330: ("PAGTO. VALE-TRANSPORTE (Qtde)", "PAGTO. VALE-TRANSPORTE (Valor)"),
    9331: (None, "PREMIO/BONUS"),
    9332: (None, "TRIBUTOS (VALOR)"),
    9333: (None, "ENCARGOS (VALOR)"),
    9334: (None, "TAXA FATURAMENTO (VALOR)"),
    9335: (None, "TAXA EXAMES MEDICOS(Valor)"),
}


# Colunas com destaque no cabeçalho da exportação (compartilhadas entre o
# caminho FEMSA atual e o renderizador por estrutura). Mesmo conteúdo de sempre.
_HEADER_DARK_RED_COLS = [
    "Nº Posicão", "CNPJ FEMSA", "Unidade - ",
    "Centro de Custo - Femsa", "Cargo - Femsa", "Motivo - Femsa"
]

_HEADER_LIGHT_RED_COLS = [
    "TAXA EXAMES MEDICOS(Valor)", "TAXA EXAMES MEDICOS(%)",
    "TAXA EXAMES MEDICOS COMPLEMENTARES (Valor)", "TAXA EXAMES MEDICOS COMPLEMENTARES (%)",
    "ENCARGOS DE FOLHA", "TAXA FATURAMENTO (VALOR)", "TAXA FATURAMENTO (%)",
    "TAXA CONTRATO (VALOR)", "TAXA CONTRATO (%)", "TRIBUTOS (VALOR)",
    "TRIBUTOS (%)", "ENCARGOS (VALOR)", "ENCARGOS (%)"
]


def normalize_name_for_match(name: str) -> str:
    """Normaliza nome para comparação (uppercase, sem espaços extras)."""
    if not name:
        return ""
    return " ".join(name.upper().strip().split())


def calcular_faturamento(df, encargos_pct=None, taxa_adm_pct=None, imposto_pct=None):
    """Conta final do faturamento por funcionário (parâmetros em %; gross-up):
      Subtotal = Total Remuneração + Encargos Sociais + Exames + Benefícios (custos)
      Taxa Adm (R$) = Subtotal × taxa_adm_pct%   -> coluna TAXA FATURAMENTO
      Total Geral   = (Subtotal + Taxa Adm) ÷ (1 − imposto_pct%)
      Tributos (R$) = Total Geral × imposto_pct% -> coluna TRIBUTOS
    Parâmetros None = não recalcula aquele pedaço (mantém o que veio).
    SUPOSIÇÕES DE MAPEAMENTO (a confirmar): taxa adm -> 'TAXA FATURAMENTO'; imposto -> 'TRIBUTOS'.
    """
    if df.empty or "Total Remuneração" not in df.columns:
        return df

    def col(c):
        return pd.to_numeric(df[c], errors="coerce").fillna(0) if c in df.columns else 0

    total_rem = col("Total Remuneração")
    # Encargos sociais (campo %); se não informado, mantém o já calculado (57,91%)
    if encargos_pct is not None:
        df["Encargos Sociais"] = (total_rem * (float(encargos_pct) / 100.0)).round(2)
        df["ENCARGOS (%)"] = encargos_pct
    encargos = col("Encargos Sociais")

    # Subtotal de custos do colaborador (gastos). Uniformes/EPIs/Equip/Treinamentos
    # entram quando as colunas existirem no df (col() retorna 0 se não existir),
    # para que os valores preenchidos pela Frente 3 fluam ao Sub-Total → taxa adm → gross-up.
    custos = (total_rem + encargos
              + col("(FAT) EXAMES MEDICOS")
              + col("PAGTO. VALE-TRANSPORTE (Valor)") + col("PAGTO. VALE REFEICAO (Valor)")
              + col("AJUDA CUSTO COMBUSTÍVEL / KM") + col("PREMIO/BONUS") + col("SEGURO DE VIDA")
              + col("UNIFORMES (Valor)") + col("EPIS (Valor)")
              + col("EQUIPAMENTOS (Valor)") + col("TREINAMENTOS (Valor)"))
    df["Sub-Total"] = custos.round(2)

    # Taxa administrativa (campo %) -> TAXA FATURAMENTO
    if taxa_adm_pct is not None:
        df["TAXA FATURAMENTO (%)"] = taxa_adm_pct
        df["TAXA FATURAMENTO (VALOR)"] = (custos * (float(taxa_adm_pct) / 100.0)).round(2)
    taxa_adm_val = col("TAXA FATURAMENTO (VALOR)")

    base = custos + taxa_adm_val
    # Valor total (gross-up pela alíquota de imposto) + impostos -> TRIBUTOS
    if imposto_pct is not None and float(imposto_pct) < 100:
        aliq = float(imposto_pct) / 100.0
        total_geral = (base / (1 - aliq)).round(2)
        df["Total Geral"] = total_geral
        df["TRIBUTOS (%)"] = imposto_pct
        df["TRIBUTOS (VALOR)"] = (total_geral * aliq).round(2)
    else:
        df["Total Geral"] = base.round(2)
    return df


def billing_to_femsa_excel(
    grouped_employees: List[Dict[str, Any]],
    periodo: str,
    codccu: str,
    cnpj_unidade: str = "15.541.957/0001-12",
    exams_data: Dict[str, float] = None,
    exams_by_numcad: Dict[int, float] = None,
    exams_by_cpf: Dict[str, float] = None,
    benefits_data: Dict[str, float] = None,
    extra_event_map: Dict[int, str] = None,
    encargos_pct: float = None,
    taxa_adm_pct: float = None,
    imposto_pct: float = None,
    colunas: List[str] = None,
    uniformes_by_numcad: Dict[int, float] = None,
    epis_by_numcad: Dict[int, float] = None,
    equipamentos_by_numcad: Dict[int, float] = None,
    treinamentos_by_numcad: Dict[int, float] = None,
    estrutura: Dict[str, Any] = None,
    salario_formula: str = None,
    campos_config: List[Dict[str, Any]] = None
) -> bytes:
    """
    Converte dados de faturamento para formato Excel FEMSA.
    Preenche (FAT) EXAMES MEDICOS casando por CPF (exams_by_cpf) com prioridade,
    e por matrícula (exams_by_numcad) como fallback. exams_data inclui por nome.
    Se benefits_data for fornecido, inclui valores de benefícios (Sodexo).
    extra_event_map: {codeve: coluna_femsa} de eventos de benefício ATIVOS
    (cadastro configurável), mesclado ao EVENT_TO_FEMSA_MAPPING como (None, coluna).
    colunas: lista ORDENADA de colunas do modelo de faturamento. Default =
    FEMSA_COLUMNS (regressão ZERO para o FEMSA de hoje). Colunas extras conhecidas
    (UNIFORMES/EPIS/EQUIPAMENTOS/TREINAMENTOS (Valor)) só aparecem se estiverem aqui.
    uniformes_by_numcad/epis_by_numcad/equipamentos_by_numcad: {numcad: valor}
    agregado dos pedidos CONFIRMADOS. Preenchem as colunas *(Valor) casando por
    matrícula. Só têm efeito em modelos que incluam essas colunas (ex.: GERAL);
    no FEMSA (sem essas colunas) não aparecem nem somam. TREINAMENTOS fica em 0.
    """
    # DIRIGIDO POR ESTRUTURA: quando o modelo tem "estrutura" (upload de Excel),
    # o df é calculado com as colunas "campo" da estrutura (fontes) e a escrita
    # da planilha segue o layout original (ver _render_por_estrutura).
    if estrutura and not colunas:
        # Import local para evitar import circular (model_structure importa
        # GERAL_COLUMNS deste módulo).
        from app.services.model_structure import derive_colunas
        colunas = derive_colunas(estrutura)
    # DIRIGIDO POR COLUNAS: a montagem usa a lista do modelo; FEMSA é o default.
    if not colunas:
        colunas = FEMSA_COLUMNS
    if exams_data is None:
        exams_data = {}
    if exams_by_numcad is None:
        exams_by_numcad = {}
    if exams_by_cpf is None:
        exams_by_cpf = {}
    if uniformes_by_numcad is None:
        uniformes_by_numcad = {}
    if epis_by_numcad is None:
        epis_by_numcad = {}
    if equipamentos_by_numcad is None:
        equipamentos_by_numcad = {}
    if treinamentos_by_numcad is None:
        treinamentos_by_numcad = {}
    # Mescla o mapeamento fixo com os eventos de benefício configuráveis (ativos).
    event_map = dict(EVENT_TO_FEMSA_MAPPING)
    if extra_event_map:
        for _cod, _col in extra_event_map.items():
            if _col:
                event_map[int(_cod)] = (None, _col)
    if benefits_data is None:
        benefits_data = {}
    # Grade "Fórmulas" (campos_config do modelo): campos com CÓDIGO próprio têm o
    # valor substituído pela soma dos eventos configurados (entra nos totais);
    # as FÓRMULAS são aplicadas depois, sobre o df (junto da fórmula de salário).
    cfg_codigo_por_campo: Dict[str, list] = {}
    for _cfg in (campos_config or []):
        _cods = [int(x) for x in str(_cfg.get("codigo") or "").split(",") if str(x).strip().isdigit()]
        if _cods and _cfg.get("campo"):
            cfg_codigo_por_campo[_cfg["campo"]] = _cods
    def format_date_br(date_val):
        if not date_val:
            return None
        if isinstance(date_val, str):
            for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%Y-%m-%dT%H:%M:%S"]:
                try:
                    dt = datetime.strptime(date_val[:10], fmt)
                    return dt.strftime("%d/%m/%Y")
                except (ValueError, TypeError):
                    continue
            return date_val
        if hasattr(date_val, 'strftime'):
            return date_val.strftime("%d/%m/%Y")
        return date_val
    
    meses = {
        "01": "Janeiro", "02": "Fevereiro", "03": "Março", "04": "Abril",
        "05": "Maio", "06": "Junho", "07": "Julho", "08": "Agosto",
        "09": "Setembro", "10": "Outubro", "11": "Novembro", "12": "Dezembro"
    }
    mes_ref = meses.get(periodo[5:7], periodo[5:7])
    
    rows = []
    for emp in grouped_employees:
        row = {col: None for col in colunas}
        
        row["Empresa"] = "Telos Consultoria Ltda"
        row["Mês Referência"] = mes_ref
        row["Nº Posicão"] = emp.get("matricula", "")
        row["CNPJ (Inserir CNPJ agência)"] = cnpj_unidade
        row["Nome"] = emp.get("nome_funcionario", "")
        row["CPF"] = emp.get("cpf", "")
        row["Função"] = emp.get("funcao", "") or emp.get("cargo", "")
        row["Unidade - "] = emp.get("nomccu", "") or emp.get("codccu", "")
        row["Centro de Custo - Femsa"] = emp.get("nomccu", "") or emp.get("codccu", "")
        row["Cargo - Femsa"] = emp.get("cargo", "")
        row["Salário"] = _safe_salario_cell(emp.get("salario"))
        row["Dt Admissão"] = format_date_br(emp.get("data_admissao", ""))
        dt_afastamento = format_date_br(emp.get("data_afastamento", ""))
        sitafa = emp.get("sitafa")
        sitafas_demissao = {7, 13, 29}
        sitafas_afastamento = {3, 4, 5, 6, 8, 9, 10, 19, 24}
        if dt_afastamento and dt_afastamento != "31/12/1900":
            try:
                sitafa_int = int(sitafa) if sitafa else 0
            except (ValueError, TypeError):
                sitafa_int = 0
            if sitafa_int in sitafas_demissao:
                row["Dt Demissão"] = dt_afastamento
            elif sitafa_int in sitafas_afastamento:
                row["Período afastamento"] = dt_afastamento
            else:
                row["Dt Demissão"] = dt_afastamento
        
        total_proventos = 0.0
        cfg_somas: Dict[str, float] = {}  # grade "Fórmulas": somas por campo configurado
        total_descontos = 0.0
        used_codes_per_col: dict = {}
        used_codes_total: set = set()

        def col_add(col_name: str, event_cod: int, amount: float, is_ref: bool = False) -> bool:
            col_set = used_codes_per_col.setdefault(col_name, set())
            if event_cod in col_set:
                logger.warning(
                    f"Evento duplicado ignorado: cod={event_cod} ja somado em '{col_name}' "
                    f"func={emp.get('nome_funcionario','')}"
                )
                return False
            col_set.add(event_cod)
            current = row.get(col_name) or 0
            row[col_name] = current + amount
            return True

        for ev in emp.get("eventos", []):
            cod_raw = ev.get("codigo_evento", "")
            ref = ev.get("referencia_evento", 0) or 0
            val = ev.get("valor_evento", 0) or 0
            tipeve = ev.get("tipo_evento", 0)
            
            try:
                cod = int(cod_raw) if cod_raw != "" else 0
            except (ValueError, TypeError):
                cod = 0
            
            try:
                tipeve = int(tipeve)
            except (ValueError, TypeError):
                tipeve = 0
            
            if cod not in used_codes_total:
                used_codes_total.add(cod)
                if tipeve == 3:
                    total_descontos += abs(val)
                elif tipeve in (1, 2):
                    total_proventos += abs(val)

            # Grade "Fórmulas": acumula os códigos configurados por campo
            for _campo_cfg, _cods_cfg in cfg_codigo_por_campo.items():
                if cod in _cods_cfg:
                    cfg_somas[_campo_cfg] = cfg_somas.get(_campo_cfg, 0.0) + val

            desc_evento = (ev.get("descricao_evento") or "").strip()

            if cod == 2470:
                col_add("ADIANTAMENTO QUINZENAL", cod, val)
                col_add("DESCONTO DE ADIANTAMENTO", cod, val)
                continue

            if cod in event_map:
                mapping = event_map[cod]
                if mapping is None:
                    continue
                qtde_col, valor_col = mapping
                if qtde_col:
                    if cod == 200:
                        col_set = used_codes_per_col.setdefault(qtde_col, set())
                        if cod not in col_set:
                            col_set.add(cod)
                            current_ref = row.get(qtde_col) or 0
                            if ref > current_ref:
                                row[qtde_col] = ref
                    else:
                        col_add(qtde_col, cod, ref)
                if valor_col:
                    col_add(valor_col, cod, val)
            elif "adicional noturno" in desc_evento.lower():
                col_add("ADICIONAL NOTURNO (Qtde)", cod, ref)
                col_add("ADICIONAL NOTURNO (Valor)", cod, val)
            elif "hora extra" in desc_evento.lower() or "horas extras" in desc_evento.lower():
                col_add("HORAS EXTRAS (Qtde)", cod, ref)
                col_add("HORAS EXTRAS (Valor)", cod, val)
            elif cod > 0 and val != 0:
                logger.info(f"Evento nao mapeado FEMSA: cod={cod}, desc={desc_evento}, tipeve={tipeve}, val={val}, func={emp.get('nome_funcionario','')}")
        
        # Grade "Fórmulas": campo com código próprio SUBSTITUI o mapeamento padrão
        for _campo_cfg in cfg_codigo_por_campo:
            row[_campo_cfg] = round(cfg_somas.get(_campo_cfg, 0.0), 2)

        row["SALÁRIO BRUTO "] = round(total_proventos, 2)
        row["SALÁRIO LÍQUIDO"] = round(total_proventos - total_descontos, 2)
        # SEGURO DE VIDA: se um evento de benefício ativo (ex.: 2450) alimentou a
        # coluna, o valor do EVENTO prevalece; sem evento, mantém o fixo R$ 5
        # (comportamento contratual de sempre). Antes o 5 sobrescrevia o evento.
        if not row.get("SEGURO DE VIDA"):
            row["SEGURO DE VIDA"] = 5
        row["DESC. SALDO NEGATIVO"] = row.get("DESC. SALDO NEGATIVO") or 0
        
        # (FAT) EXAMES MEDICOS: casa por CPF primeiro (unifica manual + upload),
        # cai para matrícula se não houver CPF (exames legados). Evita dupla contagem.
        matricula = emp.get("matricula")
        emp_cpf = re.sub(r"\D", "", str(emp.get("cpf") or ""))
        if emp_cpf and emp_cpf in exams_by_cpf:
            row["(FAT) EXAMES MEDICOS"] = exams_by_cpf[emp_cpf]
        elif matricula and int(matricula) in exams_by_numcad:
            row["(FAT) EXAMES MEDICOS"] = exams_by_numcad[int(matricula)]

        # UNIFORMES/EPIS/EQUIPAMENTOS (Valor): valores dos pedidos CONFIRMADOS,
        # casando por matrícula (numcad). Só grava quando a coluna existe no
        # modelo (FEMSA não a tem → não aparece nem soma). TREINAMENTOS fica em 0
        # (base ainda não existe). calcular_faturamento soma essas colunas no Sub-Total.
        numcad_int = None
        if matricula not in (None, ""):
            try:
                numcad_int = int(matricula)
            except (ValueError, TypeError):
                numcad_int = None
        if numcad_int is not None:
            if "UNIFORMES (Valor)" in row and numcad_int in uniformes_by_numcad:
                row["UNIFORMES (Valor)"] = uniformes_by_numcad[numcad_int]
            if "EPIS (Valor)" in row and numcad_int in epis_by_numcad:
                row["EPIS (Valor)"] = epis_by_numcad[numcad_int]
            if "EQUIPAMENTOS (Valor)" in row and numcad_int in equipamentos_by_numcad:
                row["EQUIPAMENTOS (Valor)"] = equipamentos_by_numcad[numcad_int]
            if "TREINAMENTOS (Valor)" in row and numcad_int in treinamentos_by_numcad:
                row["TREINAMENTOS (Valor)"] = treinamentos_by_numcad[numcad_int]
        if "TREINAMENTOS (Valor)" in row and row.get("TREINAMENTOS (Valor)") is None:
            row["TREINAMENTOS (Valor)"] = 0

        nome_func = emp.get("nome_funcionario", "")
        nome_norm = normalize_name_for_match(nome_func)
        if nome_norm in exams_data:
            row["TAXA EXAMES MEDICOS(Valor)"] = exams_data[nome_norm]
        
        # Buscar valor de benefícios pelo nome do funcionário
        if nome_norm in benefits_data:
            row["VALE REFEICAO(Valor)"] = benefits_data[nome_norm]
        
        rows.append(row)
    
    if not rows:
        df = pd.DataFrame({"Mensagem": ["Nenhum dado encontrado"]})
    else:
        df = pd.DataFrame(rows)
        df = calcular_totais_remuneracao(df)
        # Fórmula de salário do MODELO (metodologia do cliente, ex.: Skyrail):
        # substitui o campo "Salário" pelo valor calculado por funcionário.
        # Fórmula inválida/erro => mantém o salário-base (avaliador nunca levanta).
        if salario_formula:
            from app.services.formula_salario import avaliar_formula as _avaliar_sal

            def _sal_calc(r):
                v = _avaliar_sal(salario_formula, {
                    "salario": r.get("Salário"),
                    "total_remuneracao": r.get("Total Remuneração"),
                    "salario_dia_qtde": r.get("SALARIO DIA (Qtde)"),
                    "dias_mes": 30,
                })
                return v if v is not None else r.get("Salário")
            df["Salário"] = df.apply(_sal_calc, axis=1)
        # Grade "Fórmulas": aplica as FÓRMULAS por campo (a variável `valor` é o
        # valor-base do campo — soma dos códigos configurados, ou o padrão).
        # Fórmula só-número = valor fixo. Erro na avaliação => mantém o valor-base.
        if campos_config:
            from app.services.formula_salario import avaliar_formula as _avaliar_cfg
            for _cfg in campos_config:
                _campo, _formula = _cfg.get("campo"), (_cfg.get("formula") or "").strip()
                if not _campo or not _formula:
                    continue
                if _campo not in df.columns:
                    df[_campo] = 0

                def _calc_campo(r, _f=_formula, _c=_campo):
                    v = _avaliar_cfg(_f, {
                        "valor": r.get(_c),
                        "salario": r.get("Salário"),
                        "total_remuneracao": r.get("Total Remuneração"),
                        "salario_dia_qtde": r.get("SALARIO DIA (Qtde)"),
                        "dias_mes": 30,
                    })
                    return v if v is not None else r.get(_c)
                df[_campo] = df.apply(_calc_campo, axis=1)
        df = calcular_faturamento(df, encargos_pct, taxa_adm_pct, imposto_pct)
        # Reindexa na ordem do modelo (garante todas as colunas do modelo, na ordem,
        # e descarta qualquer coluna auxiliar que os cálculos tenham criado fora dele).
        for _c in colunas:
            if _c not in df.columns:
                df[_c] = None
        df = df[colunas]

    # CAMINHO POR ESTRUTURA: mesmos dados/derivações do df acima, mas a escrita
    # da planilha segue o layout da planilha-modelo (CONTRATO C1). Sem estrutura,
    # o caminho atual (dirigido por colunas / FEMSA) segue INTOCADO abaixo.
    if estrutura:
        return _render_por_estrutura(df, estrutura, mes_ref)

    output = BytesIO()
    
    dark_red_cols = _HEADER_DARK_RED_COLS

    light_red_cols = _HEADER_LIGHT_RED_COLS

    dark_red_fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
    light_red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)
    dark_font = Font(color="000000", bold=True)
    
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        sheet_name = f"Faturamento {mes_ref}"[:31]
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        
        worksheet = writer.sheets[sheet_name]
        
        for idx, col in enumerate(df.columns):
            if idx < 26:
                col_letter = chr(65 + idx)
            elif idx < 702:
                col_letter = chr(64 + idx // 26) + chr(65 + idx % 26)
            else:
                col_letter = "A"
            
            cell = worksheet[f"{col_letter}1"]
            if col in dark_red_cols:
                cell.fill = dark_red_fill
                cell.font = white_font
            elif col in light_red_cols:
                cell.fill = light_red_fill
                cell.font = dark_font
            
            max_length = max(
                int(df[col].fillna('').astype(str).apply(len).max()) if len(df) > 0 else 0,
                len(str(col))
            ) + 2
            worksheet.column_dimensions[col_letter].width = min(max_length, 25)
    
    output.seek(0)
    return output.getvalue()


def _valor_celula_estrutura(v: Any) -> Any:
    """Converte valor do df (numpy/pandas) em escalar aceito pelo openpyxl."""
    if v is None:
        return None
    try:
        if pd.isna(v):
            return None
    except (TypeError, ValueError):
        pass
    if hasattr(v, "item"):
        try:
            return v.item()
        except (ValueError, AttributeError):
            pass
    return v


def _coerce_numero(v: Any) -> Any:
    """Converte uma constante de texto que representa número (ex.: "0.07",
    "0,07", "1.234,56") no float correspondente, para o Excel tratar como
    NÚMERO — assim a célula exibe no separador do idioma (0,07 em pt-BR) e as
    fórmulas que a referenciam calculam em vez de devolver #VALOR!.

    Só converte quando o texto é claramente numérico; caso contrário devolve o
    valor original (constantes de texto legítimas, ex.: rótulos, ficam intactas).
    """
    if not isinstance(v, str):
        return v
    s = v.strip()
    if not s:
        return v
    tem_ponto, tem_virgula = "." in s, "," in s
    if tem_ponto and tem_virgula:
        # formato pt-BR: ponto = milhar, vírgula = decimal (1.234,56)
        candidato = s.replace(".", "").replace(",", ".")
    elif tem_virgula:
        candidato = s.replace(",", ".")
    else:
        candidato = s
    if re.fullmatch(r"-?\d+(\.\d+)?", candidato):
        try:
            return float(candidato)
        except ValueError:
            return v
    return v


def _formato_numero_coluna(col: Dict[str, Any]) -> Optional[str]:
    """
    Formato numérico básico da coluna, inferido pela fonte/cabeçalho (consistente
    com as colunas conhecidas: (%) percentual, (Qtde) quantidade, valores em moeda).
    Datas já chegam formatadas como texto dd/mm/aaaa e não precisam de formato.
    """
    ref = " ".join(
        str(x) for x in [col.get("fonte") or "", col.get("header") or ""]
    ).upper()
    if not ref.strip():
        return None
    if "(%)" in ref:
        return "0.00"
    if "(QTDE)" in ref:
        return "#,##0.00"
    palavras_moeda = (
        "(VALOR)", "SALARIO", "SALÁRIO", "TOTAL", "SUB-TOTAL", "ENCARGOS",
        "TRIBUTOS", "TAXA", "SEGURO", "PREMIO", "BONUS", "PENSÃO", "PENSAO",
        "EXAMES", "ADIANTAMENTO", "REEMB", "DESCONTO", "SALDO",
    )
    if any(p in ref for p in palavras_moeda):
        return "#,##0.00"
    return None


def _render_por_estrutura(df: "pd.DataFrame", estrutura: Dict[str, Any], mes_ref: str) -> bytes:
    """
    Escreve a planilha seguindo a "estrutura" do modelo (CONTRATO C1 em
    app/services/model_structure.py):
    - cabeçalhos nas células/linhas indicadas em headers/header_rows;
    - dados a partir de data_row, uma linha por linha do df;
    - coluna "campo" puxa a série do df pela "fonte"; "formula" escreve o
      template com {row} substituído pela linha corrente; "constante" repete
      o valor; "vazio" fica em branco.
    """
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    aba = str(estrutura.get("aba") or "").strip() or f"Faturamento {mes_ref}"
    ws.title = aba[:31]

    headers = estrutura.get("headers") or {}
    header_rows = [r for r in (estrutura.get("header_rows") or []) if isinstance(r, int)]
    try:
        data_row = int(estrutura.get("data_row"))
    except (TypeError, ValueError):
        data_row = (max(header_rows) + 1) if header_rows else 2
    if data_row < 1:
        data_row = 1
    colunas_estrutura = [c for c in (estrutura.get("colunas") or []) if isinstance(c, dict)]

    dark_red_fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
    light_red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True)
    dark_font = Font(color="000000", bold=True)
    header_font = Font(bold=True)

    # Cabeçalhos: escreve os textos nas células indicadas, em negrito, com os
    # mesmos destaques do caminho atual quando o texto for de coluna destacada.
    for letra, por_linha in headers.items():
        if not isinstance(por_linha, dict):
            continue
        for linha, texto in por_linha.items():
            try:
                r = int(linha)
            except (TypeError, ValueError):
                continue
            if r < 1 or r >= data_row:
                continue
            cell = ws[f"{letra}{r}"]
            cell.value = texto
            if texto in _HEADER_DARK_RED_COLS:
                cell.fill = dark_red_fill
                cell.font = white_font
            elif texto in _HEADER_LIGHT_RED_COLS:
                cell.fill = light_red_fill
                cell.font = dark_font
            else:
                cell.font = header_font

    # Dados: df "Mensagem" (nenhum funcionário) => só cabeçalhos, sem linhas.
    n_linhas = 0 if "Mensagem" in df.columns else len(df)

    for col in colunas_estrutura:
        letra = col.get("letra")
        tipo = col.get("tipo")
        if not letra or not tipo:
            continue
        fonte = col.get("fonte")
        serie = df[fonte] if (tipo == "campo" and fonte in df.columns) else None
        template = str(col.get("template") or "") if tipo == "formula" else ""
        num_fmt = _formato_numero_coluna(col)

        for i in range(n_linhas):
            r = data_row + i
            cell = ws[f"{letra}{r}"]
            if tipo == "campo":
                if serie is not None:
                    cell.value = _valor_celula_estrutura(serie.iloc[i])
            elif tipo == "formula":
                if template:
                    cell.value = template.replace("{row}", str(r))
            elif tipo == "constante":
                cell.value = _coerce_numero(_valor_celula_estrutura(col.get("valor")))
            else:
                continue  # "vazio" (ou tipo desconhecido): célula em branco
            if num_fmt and (tipo == "formula" or isinstance(cell.value, (int, float))):
                cell.number_format = num_fmt

        # Largura básica pela maior linha do cabeçalho da coluna.
        textos_col = headers.get(letra) or {}
        maior = max((len(str(t)) for t in textos_col.values()), default=0)
        maior = max(maior, len(str(col.get("header") or "")))
        ws.column_dimensions[letra].width = min(max(maior + 2, 12), 25)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def generate_femsa_filename(periodo: str, codccu: str) -> str:
    meses = {
        "01": "Janeiro", "02": "Fevereiro", "03": "Marco", "04": "Abril",
        "05": "Maio", "06": "Junho", "07": "Julho", "08": "Agosto",
        "09": "Setembro", "10": "Outubro", "11": "Novembro", "12": "Dezembro"
    }
    mes = meses.get(periodo[5:7], periodo[5:7])
    ano = periodo[:4]
    ccu_safe = codccu.replace(" ", "_").replace("/", "_")
    return f"Faturamento_Consolidado_{mes}.{ano}_TELOS_{ccu_safe}.xlsx"


def payroll_to_senior_excel_bytes(
    grouped_data: List[Dict[str, Any]],
    periodo: str,
) -> bytes:
    """
    Gera Excel dinamico com todos os eventos da folha Senior.
    Colunas fixas: dados do funcionario.
    Colunas dinamicas: um par (Qtde, Valor) por evento unico encontrado nos dados.
    """
    from openpyxl.styles import Alignment
    from datetime import date as _date

    period_dt = datetime.strptime(periodo[:10], "%Y-%m-%d")
    cutoff_month = period_dt.month - 2
    cutoff_year = period_dt.year
    if cutoff_month <= 0:
        cutoff_month += 12
        cutoff_year -= 1
    cutoff = _date(cutoff_year, cutoff_month, 1)

    filtered_data = []
    for emp in grouped_data:
        data_afa = (emp.get("data_afastamento") or "").strip()
        if data_afa and data_afa != "31/12/1900":
            try:
                afa_date = datetime.strptime(data_afa, "%d/%m/%Y").date()
                if afa_date < cutoff:
                    logger.info(
                        f"Funcionario removido da Folha Senior (demitido em {data_afa}, "
                        f"cutoff {cutoff}): {emp.get('nome_funcionario')}"
                    )
                    continue
            except (ValueError, TypeError):
                pass
        filtered_data.append(emp)

    events_map: Dict[int, str] = {}
    for emp in filtered_data:
        for ev in emp.get("eventos", []):
            cod = ev.get("codigo_evento") or 0
            desc = (ev.get("descricao_evento") or "").strip()
            if cod and cod not in events_map:
                events_map[cod] = desc or str(cod)

    sorted_events = sorted(events_map.items(), key=lambda x: x[0])

    rows = []
    for emp in filtered_data:
        row: Dict[str, Any] = {
            "Matrícula": emp.get("matricula"),
            "Nome": emp.get("nome_funcionario"),
            "CPF": emp.get("cpf"),
            "Centro de Custo": emp.get("codccu"),
            "Nome CCU": emp.get("nomccu"),
            "Cargo": emp.get("cargo"),
            "Salário": _safe_salario_cell(emp.get("salario")),
            "Admissão": emp.get("data_admissao"),
            "Demissão": emp.get("data_afastamento") or "",
            "Situação": emp.get("situacao"),
        }

        for cod, desc in sorted_events:
            row[f"{cod} - {desc} (Qtde)"] = 0.0
            row[f"{cod} - {desc} (Valor)"] = 0.0

        seen_codes: set = set()
        for ev in emp.get("eventos", []):
            cod = ev.get("codigo_evento") or 0
            if not cod or cod in seen_codes:
                continue
            seen_codes.add(cod)
            desc = events_map.get(cod, str(cod))
            ref = ev.get("referencia_evento") or 0.0
            val = ev.get("valor_evento") or 0.0
            row[f"{cod} - {desc} (Qtde)"] = ref
            row[f"{cod} - {desc} (Valor)"] = val

        rows.append(row)

    df = pd.DataFrame(rows) if rows else pd.DataFrame()

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Folha Senior")
        ws = writer.sheets["Folha Senior"]

        header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
        header_font = Font(bold=True, color="D4A84B")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        fixed_cols = 10
        event_fill_qtde = PatternFill(start_color="1E2A3A", end_color="1E2A3A", fill_type="solid")
        event_fill_val = PatternFill(start_color="162030", end_color="162030", fill_type="solid")

        for col_idx, cell in enumerate(ws[1], start=1):
            if col_idx > fixed_cols:
                col_label = str(cell.value or "")
                if col_label.endswith("(Qtde)"):
                    for data_cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
                        for c in data_cell:
                            c.fill = event_fill_qtde
                elif col_label.endswith("(Valor)"):
                    for data_cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
                        for c in data_cell:
                            c.fill = event_fill_val

        for col in ws.columns:
            max_len = max(
                (len(str(cell.value or "")) for cell in col),
                default=0,
            )
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 45)

    output.seek(0)
    return output.getvalue()


def generate_senior_filename(periodo: str, codccu: str) -> str:
    meses = {
        "01": "Janeiro", "02": "Fevereiro", "03": "Marco", "04": "Abril",
        "05": "Maio", "06": "Junho", "07": "Julho", "08": "Agosto",
        "09": "Setembro", "10": "Outubro", "11": "Novembro", "12": "Dezembro"
    }
    mes = meses.get(periodo[5:7], periodo[5:7])
    ano = periodo[:4]
    ccu_safe = codccu.replace(" ", "_").replace("/", "_")
    return f"Folha_Senior_{mes}.{ano}_{ccu_safe}.xlsx"
