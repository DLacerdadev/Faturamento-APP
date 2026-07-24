"""
Parser de planilha-modelo (upload de Excel) para modelos de faturamento.

Lê um arquivo .xlsx enviado pelo usuário e extrai a "estrutura" do modelo:
onde ficam os cabeçalhos, onde começam os dados e o que cada coluna significa
(campo do sistema, fórmula, constante ou vazio). Essa estrutura é salva em
BillingModel.estrutura (JSON) e usada por excel_export.billing_to_femsa_excel
para renderizar a exportação exatamente no layout da planilha original.

CONTRATO C1 — formato do JSON "estrutura" (dono: este módulo):

{
  "aba": "<nome da sheet principal>",
  "header_rows": [1, 2],                     # linhas de cabeçalho (1-based)
  "headers": {                               # textos do cabeçalho por célula
    "A": {"1": "Empresa"},
    "B": {"1": "HORAS EXTRAS", "2": "(Qtde)"}
  },
  "data_row": 3,                             # primeira linha de dados (1-based)
  "colunas": [                               # uma entrada por coluna usada, em ordem
    {"letra": "A", "header": "Empresa",  "tipo": "campo",     "fonte": "Empresa"},
    {"letra": "B", "header": "(Qtde)",   "tipo": "formula",   "template": "=C{row}+D{row}"},
    {"letra": "C", "header": "Taxa",     "tipo": "constante", "valor": 5},
    {"letra": "D", "header": "Obs",      "tipo": "vazio"}
  ]
}

Tipos de coluna:
- "campo":     dados do sistema; "fonte" é o nome CANÔNICO de uma coluna
               conhecida (GERAL_COLUMNS de excel_export = FEMSA_COLUMNS +
               UNIFORMES/EPIS/EQUIPAMENTOS/TREINAMENTOS (Valor)). O casamento
               do cabeçalho é feito por texto normalizado (caixa alta, sem
               acento, espaços colapsados).
- "formula":   fórmula Excel parametrizada; na exportação, "{row}" é trocado
               pelo número da linha corrente. Referências à própria linha de
               dados do modelo viram {row}; referências absolutas ($X$n) ou a
               outras linhas ficam literais.
- "constante": "valor" fixo repetido em todas as linhas de dados.
- "vazio":     a célula fica em branco (coluna existe só no cabeçalho).

BillingModel.estrutura = NULL  =>  modelo dirigido por "colunas" (fluxo atual,
sem nenhuma mudança; regressão zero na exportação FEMSA).
"""
import re
import colorsys
import unicodedata
from datetime import datetime, date
from io import BytesIO
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from app.services.excel_export import GERAL_COLUMNS

# Índice de cor de tema (atributo theme="N" nas células) -> nome no clrScheme.
# Ordem oficial do OOXML (com a troca dk1/lt1): 0=fundo1, 1=texto1, 2=fundo2, ...
_THEME_NAMES = ["lt1", "dk1", "lt2", "dk2", "accent1", "accent2",
                "accent3", "accent4", "accent5", "accent6", "hlink", "folHlink"]


def _theme_palette(wb) -> Dict[str, str]:
    """Extrai a paleta de cores do tema do workbook: nome (dk1/lt1/accentN...) -> hex."""
    tema = getattr(wb, "loaded_theme", None)
    if not tema:
        return {}
    if isinstance(tema, bytes):
        tema = tema.decode("utf-8", "ignore")
    m = re.search(r"<a:clrScheme.*?</a:clrScheme>", tema, re.S)
    if not m:
        return {}
    bloco = m.group(0)
    palette: Dict[str, str] = {}
    for nome in ("dk1", "lt1", "dk2", "lt2", "accent1", "accent2",
                 "accent3", "accent4", "accent5", "accent6", "hlink", "folHlink"):
        seg = re.search(r"<a:" + nome + r">(.*?)</a:" + nome + r">", bloco, re.S)
        if not seg:
            continue
        srgb = re.search(r'srgbClr val="([0-9A-Fa-f]{6})"', seg.group(1))
        sysc = re.search(r'sysClr[^>]*lastClr="([0-9A-Fa-f]{6})"', seg.group(1))
        if srgb:
            palette[nome] = srgb.group(1).upper()
        elif sysc:
            palette[nome] = sysc.group(1).upper()
    return palette


def _aplicar_tint(hex_rgb: str, tint: float) -> str:
    """Aplica o 'tint' do Excel (clareia se >0, escurece se <0) sobre um hex."""
    if not tint:
        return hex_rgb
    r, g, b = int(hex_rgb[0:2], 16), int(hex_rgb[2:4], 16), int(hex_rgb[4:6], 16)
    h, l, s = colorsys.rgb_to_hls(r / 255.0, g / 255.0, b / 255.0)
    if tint < 0:
        l = l * (1.0 + tint)
    else:
        l = l * (1.0 - tint) + tint  # aproximação padrão: L + tint*(1-L)
    l = max(0.0, min(1.0, l))
    r, g, b = colorsys.hls_to_rgb(h, l, s)
    return "%02X%02X%02X" % (round(r * 255), round(g * 255), round(b * 255))


def _cor_para_hex(color, palette: Dict[str, str]) -> Optional[str]:
    """Resolve uma cor openpyxl (rgb/tema/indexada) para hex 'RRGGBB', ou None."""
    if color is None:
        return None
    try:
        tipo = getattr(color, "type", None)
        if tipo == "rgb":
            rgb = color.rgb
            if isinstance(rgb, str) and len(rgb) >= 6:
                return rgb[-6:].upper()
            return None
        if tipo == "theme":
            nome = _THEME_NAMES[color.theme] if 0 <= color.theme < len(_THEME_NAMES) else None
            base = palette.get(nome) if nome else None
            if not base:
                return None
            return _aplicar_tint(base, float(getattr(color, "tint", 0.0) or 0.0))
    except (AttributeError, ValueError, TypeError):
        return None
    return None


def _estilo_da_celula(cell, palette: Dict[str, str]) -> Optional[Dict[str, Any]]:
    """Captura estilo visível de uma célula: preenchimento, negrito, cor/tamanho
    da fonte — como hex concreto (independe do tema do arquivo gerado)."""
    fill_hex = None
    try:
        if cell.fill and cell.fill.patternType == "solid":
            fill_hex = _cor_para_hex(cell.fill.fgColor, palette)
    except AttributeError:
        pass
    font = cell.font
    est: Dict[str, Any] = {}
    if fill_hex:
        est["fill"] = fill_hex
    if font is not None:
        if font.bold:
            est["bold"] = True
        fc = _cor_para_hex(font.color, palette)
        if fc:
            est["font_color"] = fc
        if font.sz:
            est["size"] = float(font.sz)
    try:
        al = cell.alignment
        if al is not None:
            if al.horizontal:
                est["halign"] = al.horizontal
            if al.vertical:
                est["valign"] = al.vertical
            if al.wrap_text:
                est["wrap"] = True
    except AttributeError:
        pass
    return est or None

# Superconjunto de colunas conhecidas do sistema (modelo GERAL):
# FEMSA_COLUMNS + UNIFORMES/EPIS/EQUIPAMENTOS/TREINAMENTOS (Valor).
COLUNAS_CONHECIDAS: List[str] = list(GERAL_COLUMNS)

# Limite de linhas varridas na detecção do cabeçalho/primeira linha de dados.
_MAX_LINHAS_VARREDURA = 40


def _normalizar_texto(texto: Any) -> str:
    """Normaliza texto para casamento: caixa alta, sem acento, espaços colapsados."""
    if texto is None:
        return ""
    s = unicodedata.normalize("NFKD", str(texto))
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    return " ".join(s.upper().split())


# Mapa texto normalizado -> nome canônico da coluna conhecida.
_MAPA_CONHECIDAS: Dict[str, str] = {_normalizar_texto(c): c for c in COLUNAS_CONHECIDAS}

# Sinônimos comuns em planilhas de clientes -> coluna canônica do sistema.
# Conservador: só apelidos inequívocos (layouts reais: FEMSA, Skyrail).
_SINONIMOS: Dict[str, str] = {
    "COLABORADOR": "Nome",
    "FUNCIONARIO": "Nome",
    "NOME COMPLETO": "Nome",
    "MATRICULA": "Nº Posicão",
    "ADMISSAO": "Dt Admissão",
    "DT ADMISSAO": "Dt Admissão",
    "DATA ADMISSAO": "Dt Admissão",
    "DATA DE ADMISSAO": "Dt Admissão",
    "DEMISSAO": "Dt Demissão",
    "DT DEMISSAO": "Dt Demissão",
    "DATA DEMISSAO": "Dt Demissão",
    "CARGO": "Função",
    "ENC. SOCIAIS": "Encargos Sociais",
    "ENC SOCIAIS": "Encargos Sociais",
    "ENCARGOS": "Encargos Sociais",
    "SEG. DE VIDA": "SEGURO DE VIDA",
    "SEG DE VIDA": "SEGURO DE VIDA",
    "SEGURO VIDA": "SEGURO DE VIDA",
    "UNIFORMES": "UNIFORMES (Valor)",
    "EPIS": "EPIS (Valor)",
    "EQUIPAMENTOS": "EQUIPAMENTOS (Valor)",
    "TREINAMENTOS": "TREINAMENTOS (Valor)",
    "VT": "PAGTO. VALE-TRANSPORTE (Valor)",
    "VALE TRANSPORTE": "PAGTO. VALE-TRANSPORTE (Valor)",
    "VR": "PAGTO. VALE REFEICAO (Valor)",
    "VALE REFEICAO": "PAGTO. VALE REFEICAO (Valor)",
    "TOTAL MENSAL UNITARIO": "Total Geral",
    "TOTAL MENSAL": "Total Geral",
}


def coluna_conhecida(texto: Any) -> Optional[str]:
    """Retorna o nome CANÔNICO da coluna do sistema que casa com o texto (nome
    exato normalizado ou sinônimo conhecido), ou None."""
    norm = _normalizar_texto(texto)
    if not norm:
        return None
    return _MAPA_CONHECIDAS.get(norm) or _SINONIMOS.get(norm)


def _parametrizar_formula(formula: str, data_row: int) -> str:
    """
    Converte uma fórmula lida da primeira linha de dados em template com {row}.

    Referências relativas de linha que apontam para a PRÓPRIA linha de dados
    (ex.: C3, $C3 quando data_row=3) viram C{row}/$C{row}. Referências com
    linha absoluta ($C$3) ou apontando para outras linhas ficam literais.
    """
    padrao = re.compile(r"(\$?)([A-Za-z]{1,3})(\$?)(\d+)")

    def _sub(m: "re.Match") -> str:
        abs_col, letra, abs_row, num = m.group(1), m.group(2), m.group(3), m.group(4)
        if abs_row == "" and int(num) == data_row:
            return f"{abs_col}{letra}" + "{row}"
        return m.group(0)

    return padrao.sub(_sub, formula)


def _celula_vazia(valor: Any) -> bool:
    return valor is None or (isinstance(valor, str) and valor.strip() == "")


def _valor_json(valor: Any) -> Any:
    """Converte o valor da célula em algo serializável em JSON."""
    if isinstance(valor, datetime):
        return valor.strftime("%d/%m/%Y")
    if isinstance(valor, date):
        return valor.strftime("%d/%m/%Y")
    if isinstance(valor, str):
        return valor.strip()
    return valor


def _sheet_tem_dados(ws) -> bool:
    """Aba tem algum conteúdo nas primeiras linhas varridas?"""
    max_r = min(ws.max_row or 0, _MAX_LINHAS_VARREDURA)
    max_c = ws.max_column or 0
    for r in range(1, max_r + 1):
        for c in range(1, max_c + 1):
            if not _celula_vazia(ws.cell(row=r, column=c).value):
                return True
    return False


def _eh_linha_cabecalho(valores_nao_vazios: List[Any]) -> bool:
    """
    Linha com cara de cabeçalho: maioria de células texto (sem fórmulas).
    Linhas com fórmulas ou maioria numérica/data têm cara de dados.
    """
    textos = 0
    for v in valores_nao_vazios:
        if isinstance(v, str):
            if v.startswith("="):
                return False  # fórmula => linha de dados
        else:
            continue
        textos += 1
    return textos * 2 > len(valores_nao_vazios)


def _detectar_cabecalho(ws) -> tuple:
    """
    Detecta (header_rows, data_row) por DENSIDADE: a linha-âncora do cabeçalho
    é a com MAIS células de texto nas primeiras linhas (planilhas reais têm
    blocos de título/emitente esparsos acima do cabeçalho — ex.: relatório
    Skyrail — que não podem ser confundidos com cabeçalho nem com dados).
    O bloco de cabeçalho se estende para as linhas vizinhas majoritariamente
    de texto (cabeçalhos de 2 linhas); data_row = primeira linha não-vazia
    abaixo do bloco.
    """
    max_r = min(ws.max_row or 1, _MAX_LINHAS_VARREDURA)
    max_c = ws.max_column or 1

    info = []  # (linha, n_nao_vazios, eh_texto_majoritario)
    for r in range(1, max_r + 1):
        valores = [ws.cell(row=r, column=c).value for c in range(1, max_c + 1)]
        nao_vazios = [v for v in valores if not _celula_vazia(v)]
        if not nao_vazios:
            continue
        info.append((r, len(nao_vazios), _eh_linha_cabecalho(nao_vazios)))

    if not info:
        return [1], 2

    # Âncora: linha de texto com maior densidade (mín. 3 células p/ ser cabeçalho).
    candidatas = [(r, n) for r, n, eh_txt in info if eh_txt and n >= 3]
    if not candidatas:
        # fallback: comportamento antigo (primeira linha de texto é o cabeçalho)
        header_rows = [r for r, _, eh_txt in info if eh_txt][:1] or [info[0][0]]
        depois = [r for r, _, _ in info if r > header_rows[-1]]
        return header_rows, (depois[0] if depois else header_rows[-1] + 1)

    ancora, densidade = max(candidatas, key=lambda x: x[1])

    # Estende o bloco para vizinhas de texto com densidade relevante (>= 1/3 da âncora).
    por_linha = {r: (n, eh_txt) for r, n, eh_txt in info}
    header_rows = [ancora]
    r = ancora - 1
    while r >= 1 and r in por_linha and por_linha[r][1] and por_linha[r][0] * 3 >= densidade:
        header_rows.insert(0, r)
        r -= 1
    r = ancora + 1
    while r <= max_r and r in por_linha and por_linha[r][1] and por_linha[r][0] * 3 >= densidade:
        header_rows.append(r)
        r += 1

    depois = [x for x, _, _ in info if x > header_rows[-1]]
    data_row = depois[0] if depois else header_rows[-1] + 1
    return header_rows, data_row


def _casar_fonte(textos_header: Dict[str, str]) -> Optional[str]:
    """
    Tenta casar o cabeçalho da coluna com uma coluna conhecida do sistema.
    Testa o texto mais próximo dos dados (última linha do cabeçalho) e a
    concatenação de todas as linhas (ex.: "HORAS EXTRAS" + "(Qtde)").
    """
    if not textos_header:
        return None
    linhas_ordenadas = sorted(textos_header.keys(), key=int)
    candidatos = [textos_header[linhas_ordenadas[-1]]]
    if len(linhas_ordenadas) > 1:
        candidatos.append(" ".join(textos_header[k] for k in linhas_ordenadas))
    for cand in candidatos:
        canonico = coluna_conhecida(cand)
        if canonico:
            return canonico
    return None


def parse_model_xlsx(conteudo: bytes, nome_arquivo: str = "") -> Dict[str, Any]:
    """
    Lê os bytes de um .xlsx modelo e devolve o dict "estrutura" (CONTRATO C1).

    Passos:
    1. Abre com openpyxl (data_only=False, para enxergar as fórmulas).
    2. Escolhe a aba principal: primeira aba visível com dados.
    3. Detecta linhas de cabeçalho e a primeira linha de dados.
    4. Classifica cada coluna usada olhando a PRIMEIRA linha de dados:
       fórmula -> "formula"; cabeçalho conhecido -> "campo"; célula vazia ->
       "vazio"; senão -> "constante".

    Levanta ValueError com mensagem em PT-BR se o arquivo for inválido.
    """
    try:
        wb = load_workbook(BytesIO(conteudo), data_only=False)
    except Exception as exc:
        nome = nome_arquivo or "arquivo enviado"
        raise ValueError(f"Não foi possível ler '{nome}' como Excel (.xlsx): {exc}")

    ws = None
    for sheet in wb.worksheets:
        if getattr(sheet, "sheet_state", "visible") == "visible" and _sheet_tem_dados(sheet):
            ws = sheet
            break
    if ws is None:
        raise ValueError("Nenhuma aba visível com dados foi encontrada na planilha-modelo.")

    header_rows, data_row = _detectar_cabecalho(ws)

    palette = _theme_palette(wb)  # para resolver cores de tema -> hex

    headers: Dict[str, Dict[str, str]] = {}
    estilos_header: Dict[str, Dict[str, Any]] = {}   # letra -> {linha -> estilo}
    colunas: List[Dict[str, Any]] = []
    max_col = ws.max_column or 1

    for idx in range(1, max_col + 1):
        letra = get_column_letter(idx)

        textos: Dict[str, str] = {}
        estilos_linha: Dict[str, Any] = {}
        for r in header_rows:
            cel = ws.cell(row=r, column=idx)
            if not _celula_vazia(cel.value):
                textos[str(r)] = str(cel.value).strip()
                est = _estilo_da_celula(cel, palette)
                if est:
                    estilos_linha[str(r)] = est

        celula = ws.cell(row=data_row, column=idx)
        celula_dado = celula.value

        # Coluna sem cabeçalho e sem dado na primeira linha: ignorada.
        if not textos and _celula_vazia(celula_dado):
            continue

        if textos:
            headers[letra] = textos
            if estilos_linha:
                estilos_header[letra] = estilos_linha
            linhas_ordenadas = sorted(textos.keys(), key=int)
            header_txt = textos[linhas_ordenadas[-1]]
        else:
            header_txt = ""

        col: Dict[str, Any] = {"letra": letra, "header": header_txt}

        if isinstance(celula_dado, str) and celula_dado.startswith("="):
            col["tipo"] = "formula"
            col["template"] = _parametrizar_formula(celula_dado, data_row)
        else:
            fonte = _casar_fonte(textos)
            if fonte:
                col["tipo"] = "campo"
                col["fonte"] = fonte
            elif _celula_vazia(celula_dado):
                col["tipo"] = "vazio"
            else:
                col["tipo"] = "constante"
                col["valor"] = _valor_json(celula_dado)

        # Estilo da célula de DADOS (aplicado a todas as linhas de dados da coluna).
        est_dado = _estilo_da_celula(celula, palette)
        if est_dado:
            col["estilo"] = est_dado

        colunas.append(col)

    # Bloco de TOPO: linhas acima do cabeçalho (título, emitente, logo, datas...).
    # O detector de cabeçalho aponta header_rows (rótulos das colunas); tudo que
    # vem ANTES é o cabeçalho institucional do relatório — capturado para
    # reaparecer na exportação. Guarda valor (fórmula '=...' preservada) + estilo.
    min_hdr = min(header_rows) if header_rows else data_row
    topo: Dict[str, Dict[str, Any]] = {}
    for r in range(1, min_hdr):
        for idx in range(1, max_col + 1):
            cel = ws.cell(row=r, column=idx)
            if _celula_vazia(cel.value):
                continue
            letra = get_column_letter(idx)
            v = cel.value
            val = v if (isinstance(v, str) and v.startswith("=")) else _valor_json(v)
            topo.setdefault(letra, {})[str(r)] = {"v": val, "estilo": _estilo_da_celula(cel, palette)}

    # Mesclagens (merges) inteiramente na região do cabeçalho (linhas < data_row):
    # títulos e grupos como "BENEFÍCIOS", "TAXA ADM." dependem delas para o layout.
    merges: List[str] = []
    for rng in ws.merged_cells.ranges:
        try:
            if rng.max_row < data_row:
                merges.append(str(rng))
        except AttributeError:
            continue

    return {
        "estilos_header": estilos_header,
        "topo": topo,
        "merges": merges,
        "aba": ws.title,
        "header_rows": header_rows,
        "headers": headers,
        "data_row": data_row,
        "colunas": colunas,
    }


def derive_colunas(estrutura: Dict[str, Any]) -> List[str]:
    """
    Nomes canônicos das colunas tipo "campo", na ordem da estrutura.
    Compatibilidade com o fluxo dirigido por colunas (BillingModel.colunas).
    """
    if not isinstance(estrutura, dict):
        return []
    return [
        c.get("fonte")
        for c in (estrutura.get("colunas") or [])
        if isinstance(c, dict) and c.get("tipo") == "campo" and c.get("fonte")
    ]


def validate_estrutura(estrutura: Dict[str, Any]) -> List[str]:
    """
    Valida uma estrutura (CONTRATO C1) e retorna a lista de problemas
    encontrados (vazia = estrutura utilizável). Mensagens em PT-BR.
    """
    problemas: List[str] = []
    if not isinstance(estrutura, dict):
        return ["Estrutura inválida: esperado um objeto JSON."]

    if not str(estrutura.get("aba") or "").strip():
        problemas.append("Estrutura sem nome de aba ('aba').")

    data_row = estrutura.get("data_row")
    if not isinstance(data_row, int) or data_row < 1:
        problemas.append("'data_row' deve ser um número inteiro maior ou igual a 1.")
    else:
        header_rows = estrutura.get("header_rows") or []
        for hr in header_rows:
            if isinstance(hr, int) and hr >= data_row:
                problemas.append(
                    f"Linha de cabeçalho {hr} deve ficar ACIMA da primeira linha de dados ({data_row})."
                )

    colunas = estrutura.get("colunas")
    if not isinstance(colunas, list) or not colunas:
        problemas.append("Estrutura sem colunas.")
        return problemas

    letras_vistas = set()
    tem_campo = False
    for i, col in enumerate(colunas, start=1):
        rotulo = f"#{i}"
        if not isinstance(col, dict):
            problemas.append(f"Coluna {rotulo} inválida (esperado um objeto).")
            continue

        letra = col.get("letra")
        if letra:
            rotulo = str(letra)
        if not letra or not re.fullmatch(r"[A-Z]{1,3}", str(letra)):
            problemas.append(f"Coluna {rotulo}: letra de coluna ausente ou inválida.")
        elif letra in letras_vistas:
            problemas.append(f"Coluna {rotulo}: letra duplicada na estrutura.")
        else:
            letras_vistas.add(letra)

        tipo = col.get("tipo")
        if tipo not in ("campo", "formula", "constante", "vazio"):
            problemas.append(f"Coluna {rotulo}: tipo inválido '{tipo}'.")
        elif tipo == "campo":
            fonte = col.get("fonte")
            if not fonte:
                problemas.append(f"Coluna {rotulo}: tipo 'campo' sem 'fonte'.")
            elif fonte not in COLUNAS_CONHECIDAS:
                problemas.append(f"Coluna {rotulo}: fonte desconhecida '{fonte}'.")
            else:
                tem_campo = True
        elif tipo == "formula" and not str(col.get("template") or "").strip():
            problemas.append(f"Coluna {rotulo}: tipo 'formula' sem 'template'.")

    if not tem_campo:
        problemas.append(
            "Estrutura sem nenhuma coluna do tipo 'campo' — nada seria preenchido com dados do sistema."
        )
    return problemas
