"""
Geração do Excel da Solicitação de Compra (pedido de compra — EPI, uniforme e
equipamento, inclusive pedidos MISTOS).

Layout (revisão 2026-07 — pedido centrado em ITENS, sem nomes de funcionários):
- Cabeçalho com empresa / CCU / competência / solicitante / data-hora
- Tabela de itens: Item, Categoria, Tamanho, C.A, Qtde, Valor unit., Valor total
- Linha TOTAL DO PEDIDO em destaque

Os funcionários NÃO aparecem no pedido de compra — o vínculo por matrícula
(employee_numcad) continua persistido nas linhas e é usado só pelo faturamento.
"""
from datetime import datetime
from io import BytesIO
from typing import Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.models.epi_purchase import EpiPurchasePackage, EpiPurchaseItem


GOLD_FILL = PatternFill(start_color="D4A84B", end_color="D4A84B", fill_type="solid")
DARK_FILL = PatternFill(start_color="1A1A1A", end_color="1A1A1A", fill_type="solid")
SUBTLE_FILL = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="FFF4DC", end_color="FFF4DC", fill_type="solid")
THIN = Side(border_style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CATEGORIA_LABEL = {"epi": "EPI", "uniforme": "Uniforme", "equipamento": "Equipamento"}


def _categoria_item(it: EpiPurchaseItem, pkg: EpiPurchasePackage) -> str:
    return (it.categoria or pkg.categoria or "epi").strip().lower()


def _group_items(pkg: EpiPurchasePackage) -> List[Dict]:
    """
    Agrupa as linhas (funcionário × item) na visão "por item" do pedido:
    chave = (categoria, item do catálogo, tamanho, valor_unitario).

    Quantidade do item: prefere `quantidade_total_item` (override do usuário,
    replicado nas linhas do cartesiano); sem override, soma `quantidade` das
    linhas (modo 1:1). `valor_total` = quantidade × valor_unitario.
    """
    items = list(pkg.items or [])
    buckets: Dict[Tuple, Dict] = {}
    for it in items:
        cat = _categoria_item(it, pkg)
        item_key = it.epi_id if it.epi_id is not None else ("p:" + str(it.produto_codigo or ""))
        key = (cat, item_key, it.tamanho or "", round(it.valor_unitario or 0.0, 6))
        bucket = buckets.setdefault(key, {
            "categoria": cat,
            "descricao": it.descricao or "",
            "tamanho": it.tamanho or "",
            "ca_numero": it.ca_numero or "",
            "valor_unitario": it.valor_unitario or 0.0,
            "valor_unitario_catalogo": it.valor_unitario_catalogo,
            "quantidade_override": None,
            "_quantidade_soma": 0,
        })
        if bucket["quantidade_override"] is None and it.quantidade_total_item is not None:
            bucket["quantidade_override"] = it.quantidade_total_item
        bucket["_quantidade_soma"] += (it.quantidade or 0)
        if not bucket["ca_numero"] and it.ca_numero:
            bucket["ca_numero"] = it.ca_numero

    out = []
    for b in buckets.values():
        b["quantidade"] = b["quantidade_override"] if b["quantidade_override"] is not None else b["_quantidade_soma"]
        b["valor_total"] = round((b["quantidade"] or 0) * (b["valor_unitario"] or 0.0), 2)
        del b["quantidade_override"], b["_quantidade_soma"]
        out.append(b)
    # Ordena por categoria (EPI, uniforme, equipamento) e nome do item
    ordem_cat = {"epi": 0, "uniforme": 1, "equipamento": 2}
    out.sort(key=lambda b: (ordem_cat.get(b["categoria"], 9), (b["descricao"] or "").upper(), b["tamanho"]))
    return out


def _format_money(value: float) -> str:
    return f"R$ {value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def generate_solicitacao_xlsx(pkg: EpiPurchasePackage) -> bytes:
    """Gera o Excel da solicitação de compra (pedido) como bytes.

    Uma linha POR ITEM do pedido (agrupado por categoria/item/tamanho/valor),
    sem nomes de funcionários. Total do pedido ao final.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Solicitação"

    bold = Font(name="Calibri", size=11, bold=True)
    bold_gold = Font(name="Calibri", size=11, bold=True, color="1A1A1A")
    title = Font(name="Calibri", size=14, bold=True, color="1A1A1A")
    muted = Font(name="Calibri", size=10, color="6B6B6B")

    NCOL = 7  # Item | Categoria | Tamanho | C.A | Qtde | Valor unit. | Valor total

    # Título
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NCOL)
    ws["A1"] = "SOLICITAÇÃO DE COMPRA"
    ws["A1"].font = title
    ws["A1"].fill = GOLD_FILL
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Cabeçalho
    competencia = pkg.mes_ano.strftime("%m/%Y") if pkg.mes_ano else "—"
    gen_at = (pkg.solicitacao_generated_at or datetime.utcnow()).strftime("%d/%m/%Y %H:%M")
    linhas_itens = _group_items(pkg)
    cats = sorted({b["categoria"] for b in linhas_itens}, key=lambda c: {"epi": 0, "uniforme": 1, "equipamento": 2}.get(c, 9))
    categorias_str = " + ".join(CATEGORIA_LABEL.get(c, c.capitalize()) for c in cats) or "—"

    header_rows = [
        ("Empresa solicitante", pkg.empresa or "—"),
        ("Centro de custo", pkg.codccu or "—"),
        ("Competência", competencia),
        ("Categorias", categorias_str),
        ("Solicitante", pkg.solicitante_nome or "—"),
        ("Gerada em", gen_at),
        ("ID do pedido", f"#{pkg.id}"),
    ]
    row = 3
    for label, value in header_rows:
        ws.cell(row=row, column=1, value=label).font = bold
        ws.cell(row=row, column=1).fill = SUBTLE_FILL
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=NCOL)
        ws.cell(row=row, column=2, value=value).alignment = Alignment(horizontal="left")
        row += 1

    # Tabela por item
    row += 1
    headers = ["Item", "Categoria", "Tamanho", "C.A", "Qtde", "Valor unit.", "Valor total"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = bold_gold
        cell.fill = GOLD_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
    row += 1

    total_qtde = 0
    total_valor = 0.0
    for b in linhas_itens:
        ws.cell(row=row, column=1, value=b["descricao"] or "—").border = BORDER
        ws.cell(row=row, column=2, value=CATEGORIA_LABEL.get(b["categoria"], b["categoria"].capitalize())).border = BORDER
        ws.cell(row=row, column=3, value=b["tamanho"] or "").border = BORDER
        ws.cell(row=row, column=4, value=b["ca_numero"] or "").border = BORDER
        ws.cell(row=row, column=5, value=b["quantidade"]).border = BORDER
        ws.cell(row=row, column=6, value=_format_money(b["valor_unitario"])).border = BORDER
        ws.cell(row=row, column=7, value=_format_money(b["valor_total"])).border = BORDER

        # Aviso visual se valor difere do catálogo
        if b["valor_unitario_catalogo"] is not None and abs((b["valor_unitario_catalogo"] or 0) - (b["valor_unitario"] or 0)) > 0.001:
            ws.cell(row=row, column=6).font = Font(name="Calibri", size=11, italic=True, color="B8923F")

        total_qtde += b["quantidade"] or 0
        total_valor += b["valor_total"] or 0.0
        row += 1

    # Total do pedido
    for col in range(1, NCOL + 1):
        ws.cell(row=row, column=col).fill = TOTAL_FILL
        ws.cell(row=row, column=col).border = BORDER
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.cell(row=row, column=1, value="TOTAL DO PEDIDO").font = bold
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="right")
    ws.cell(row=row, column=5, value=total_qtde).font = bold
    ws.cell(row=row, column=7, value=_format_money(total_valor)).font = bold

    # Larguras
    widths = [42, 14, 10, 12, 8, 14, 14]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    # Rodapé
    row += 2
    ws.cell(row=row, column=1, value="Gerado automaticamente pelo Faturamento App — Telos Consultoria").font = muted

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_filename(pkg: EpiPurchasePackage) -> str:
    """Retorna o nome canônico do arquivo da solicitação."""
    ts = (pkg.solicitacao_generated_at or datetime.utcnow()).strftime("%Y%m%d-%H%M%S")
    return f"pedido_compra_{pkg.id}_{ts}.xlsx"
